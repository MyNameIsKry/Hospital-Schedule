import random
import copy
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from collections import defaultdict

# =====================================================
# GLOBAL CONFIG
# =====================================================
NUM_DAYS = 30

# Cấu trúc: Mỗi khoa có nhiều phòng
DEPARTMENTS = {
    "Nội khoa": ["Phòng Nội 1", "Phòng Nội 2"],
    "Ngoại khoa": ["Phòng Ngoại 1", "Phòng Ngoại 2"],
    "Sản khoa": ["Phòng Sản 1", "Phòng Sản 2"],
    "Nhi khoa": ["Phòng Nhi 1", "Phòng Nhi 2"],
    "Tim mạch": ["Phòng Tim 1", "Phòng Tim 2"]
}

# =====================================================
# STAFF CONFIGURATION
# =====================================================
DOCTORS_PER_DEPARTMENT = 25
NURSES_PER_DEPARTMENT = 40
SENIOR_DOCTOR_RATIO = 0.3
SENIOR_NURSE_RATIO = 0.25

# shifts: sáng – chiều – tối
SHIFTS = [
    ("Sáng", 6, 12, 6),
    ("Chiều", 12, 18, 6),
    ("Tối", 18, 30, 12)
]

# ---------------- HARD CONSTRAINTS (MUST SATISFY) ----------------
MIN_DOCTOR_PER_SHIFT = 2
MIN_NURSE_PER_SHIFT  = 3
MIN_TOTAL_PER_SHIFT  = 5
MIN_EXPERIENCE_YEARS = 5

# ---------------- SOFT CONSTRAINTS (SHOULD SATISFY) ----------------
MAX_HOURS_PER_WEEK   = 30      # <= 30h/tuần
MIN_REST_HOURS       = 12      # >= 12h giữa 2 ca
MAX_HOURS_PER_MONTH = 120      # <= 120h/tháng
MIN_HOURS_PER_MONTH = 60       # >= 60h/tháng

# ---------------- GA CONFIG ----------------
POPULATION_SIZE = 100
GENERATIONS = 150
ELITE_SIZE = 10
TOURNAMENT_K = 10
PARENT_POOL_RATIO = 0.25
MUTATION_RATE = 0.15

STAGNATION_LIMIT = 5
HILL_CLIMB_STEPS = 50

# ---------------- PENALTY WEIGHTS ----------------
# HARD CONSTRAINTS - Phạt cực nặng (không được vi phạm)
W_NO_DOCTOR   = 1_000_000
W_NO_NURSE    = 1_000_000
W_LESS_5      = 800_000
W_NO_SENIOR   = 1_200_000
W_WRONG_DEPT  = 1_500_000
W_DAY_OFF     = 2_000_000

# SOFT CONSTRAINTS - Phạt nhẹ hơn (có thể vi phạm một chút)
W_OVER_30H    = 500      # Giảm từ 50,000 → 500
W_NO_REST     = 800      # Giảm từ 80,000 → 800
W_OVER_MONTHLY = 300     # Giảm từ 30,000 → 300
W_UNDER_MONTHLY = 200    # Giảm từ 20,000 → 200
W_FAIRNESS    = 5


class Employee:
    def __init__(self, id, name, role, department, days_off, years_exp):
        self.id = id
        self.name = name
        self.role = role
        self.department = department
        self.days_off = set(days_off)
        self.years_exp = years_exp

class Shift:
    def __init__(self, name, start, end, hours):
        self.name = name
        self.start = start
        self.end = end
        self.hours = hours


def generate_sample_data():
    random.seed(42)
    
    shifts = [Shift(*s) for s in SHIFTS]
    days = list(range(NUM_DAYS))
    
    all_rooms = []
    dept_to_rooms = {}
    for dept, rooms in DEPARTMENTS.items():
        all_rooms.extend(rooms)
        dept_to_rooms[dept] = rooms
    
    employees = []
    eid = 0
    
    dept_list = list(DEPARTMENTS.keys())
    
    for dept in dept_list:
        for i in range(DOCTORS_PER_DEPARTMENT):
            if random.random() < SENIOR_DOCTOR_RATIO:
                years_exp = random.choice([5, 6, 7, 8, 10])
            else:
                years_exp = random.randint(1, 4)
            
            employees.append(Employee(
                eid, 
                f"BS_{dept[:3]}_{i+1}", 
                "doctor",
                dept,
                random.sample(days, 1),
                years_exp
            ))
            eid += 1
    
    for dept in dept_list:
        for i in range(NURSES_PER_DEPARTMENT):
            if random.random() < SENIOR_NURSE_RATIO:
                years_exp = random.choice([5, 6, 7, 8, 9, 10])
            else:
                years_exp = random.randint(1, 4)
            
            employees.append(Employee(
                eid, 
                f"DD_{dept[:3]}_{i+1}", 
                "nurse",
                dept,
                random.sample(days, 1),
                years_exp
            ))
            eid += 1
    
    print("=" * 80)
    print("THÔNG TIN NHÂN VIÊN")
    print("=" * 80)
    print(f"Số khoa: {len(DEPARTMENTS)}")
    print(f"Tổng số phòng: {len(all_rooms)}")
    print(f"Tổng số bác sĩ: {sum(e.role == 'doctor' for e in employees)}")
    print(f"Tổng số điều dưỡng: {sum(e.role == 'nurse' for e in employees)}")
    print(f"Nhân viên có kinh nghiệm ≥5 năm: {sum(e.years_exp >= 5 for e in employees)}")
    print(f"\n STAFF CONFIGURATION:")
    print(f"   Bác sĩ mỗi khoa: {DOCTORS_PER_DEPARTMENT}")
    print(f"   Điều dưỡng mỗi khoa: {NURSES_PER_DEPARTMENT}")
    print(f"   Tỷ lệ bác sĩ senior: {SENIOR_DOCTOR_RATIO*100:.0f}%")
    print(f"   Tỷ lệ điều dưỡng senior: {SENIOR_NURSE_RATIO*100:.0f}%")
    print("\n Phân bổ theo khoa:")
    for dept in dept_list:
        dept_docs = sum(1 for e in employees if e.department == dept and e.role == 'doctor')
        dept_nurses = sum(1 for e in employees if e.department == dept and e.role == 'nurse')
        senior_docs = sum(1 for e in employees if e.department == dept and e.role == 'doctor' and e.years_exp >= 5)
        senior_nurses = sum(1 for e in employees if e.department == dept and e.role == 'nurse' and e.years_exp >= 5)
        print(f"   {dept}: {dept_docs} BS ({senior_docs} senior), {dept_nurses} DD ({senior_nurses} senior), {len(DEPARTMENTS[dept])} phòng")
    print("=" * 80 + "\n")
    
    return employees, dept_to_rooms, shifts, days


def get_room_department(room, dept_to_rooms):
    """Tìm khoa của phòng"""
    for dept, rooms in dept_to_rooms.items():
        if room in rooms:
            return dept
    return None


def create_individual(employees, dept_to_rooms, shifts, days):
    schedule = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
    
    emp_shift_count = defaultdict(int)
    emp_hours = defaultdict(int)
    
    all_rooms = [room for rooms in dept_to_rooms.values() for room in rooms]
    
    for d in days:
        for s in shifts:
            for room in all_rooms:
                dept = get_room_department(room, dept_to_rooms)
                
                pool = [
                    e for e in employees
                    if e.department == dept and d not in e.days_off
                ]
                
                doctors = [e for e in pool if e.role == "doctor"]
                nurses  = [e for e in pool if e.role == "nurse"]
                
                doctors.sort(key=lambda x: (emp_hours[x.id], emp_shift_count[x.id]))
                nurses.sort(key=lambda x: (emp_hours[x.id], emp_shift_count[x.id]))
                
                selected_doctors = []
                if len(doctors) >= MIN_DOCTOR_PER_SHIFT:
                    selected_doctors = doctors[:MIN_DOCTOR_PER_SHIFT]
                elif doctors:
                    selected_doctors = doctors
                
                selected_nurses = []
                if len(nurses) >= MIN_NURSE_PER_SHIFT:
                    selected_nurses = nurses[:MIN_NURSE_PER_SHIFT]
                elif nurses:
                    selected_nurses = nurses
                
                selected_all = selected_doctors + selected_nurses
                
                has_senior = any(e.years_exp >= MIN_EXPERIENCE_YEARS for e in selected_all)
                if not has_senior:
                    seniors = [e for e in pool if e.years_exp >= MIN_EXPERIENCE_YEARS]
                    seniors.sort(key=lambda x: (emp_hours[x.id], emp_shift_count[x.id]))
                    if seniors:
                        senior = seniors[0]
                        if senior not in selected_all:
                            selected_all.append(senior)
                
                for e in selected_all:
                    emp_shift_count[e.id] += 1
                    emp_hours[e.id] += s.hours
                
                schedule[d][s.name][room] = [e.id for e in selected_all]
    
    return schedule


def check_constraints_detailed(schedule, employees, dept_to_rooms, shifts, days):
    emp = {e.id: e for e in employees}
    
    hours_week = defaultdict(int)
    timeline = defaultdict(list)
    
    hard_violations = {
        'no_doctor': [],
        'no_nurse': [],
        'less_than_5': [],
        'no_senior': [],
        'wrong_dept': [],
        'day_off': []
    }
    
    soft_violations = {
        'over_30h': [],
        'no_rest_12h': [],
        'over_monthly': [],
        'under_monthly': []
    }
    
    soft_stats = {
        'total_hours': defaultdict(int),
        'shift_counts': defaultdict(int)
    }
    
    all_rooms = [room for rooms in dept_to_rooms.values() for room in rooms]
    
    for d in days:
        for s in shifts:
            for room in all_rooms:
                dept = get_room_department(room, dept_to_rooms)
                ids = schedule[d][s.name][room]
                
                doctors = [i for i in ids if emp[i].role == "doctor"]
                nurses  = [i for i in ids if emp[i].role == "nurse"]
                
                if len(doctors) < MIN_DOCTOR_PER_SHIFT:
                    hard_violations['no_doctor'].append({
                        'day': d + 1, 'shift': s.name, 'room': room,
                        'dept': dept, 'required': MIN_DOCTOR_PER_SHIFT,
                        'actual': len(doctors), 'missing': MIN_DOCTOR_PER_SHIFT - len(doctors)
                    })
                
                if len(nurses) < MIN_NURSE_PER_SHIFT:
                    hard_violations['no_nurse'].append({
                        'day': d + 1, 'shift': s.name, 'room': room,
                        'dept': dept, 'required': MIN_NURSE_PER_SHIFT,
                        'actual': len(nurses), 'missing': MIN_NURSE_PER_SHIFT - len(nurses)
                    })
                
                total_staff = len(doctors) + len(nurses)
                if total_staff < MIN_TOTAL_PER_SHIFT:
                    hard_violations['less_than_5'].append({
                        'day': d + 1, 'shift': s.name, 'room': room,
                        'dept': dept, 'required': MIN_TOTAL_PER_SHIFT,
                        'actual': total_staff, 'missing': MIN_TOTAL_PER_SHIFT - total_staff
                    })
                
                assigned_emps = [emp[i] for i in ids]
                has_senior = any(e.years_exp >= MIN_EXPERIENCE_YEARS for e in assigned_emps)
                if not has_senior:
                    hard_violations['no_senior'].append({
                        'day': d + 1, 'shift': s.name, 'room': room,
                        'dept': dept, 'staff': [f"{e.name}({e.years_exp}y)" for e in assigned_emps]
                    })
                
                for i in ids:
                    if emp[i].department != dept:
                        hard_violations['wrong_dept'].append({
                            'day': d + 1, 'shift': s.name,
                            'employee': emp[i].name,
                            'assigned_dept': dept,
                            'correct_dept': emp[i].department,
                            'room': room
                        })
                    
                    if d in emp[i].days_off:
                        hard_violations['day_off'].append({
                            'day': d + 1, 'shift': s.name,
                            'room': room, 'employee': emp[i].name
                        })
                    
                    week = d // 7
                    hours_week[(i, week)] += s.hours
                    timeline[i].append((d, s))
                    soft_stats['total_hours'][i] += s.hours
                    soft_stats['shift_counts'][i] += 1
    
    # Kiểm tra ràng buộc mềm: giờ làm/tuần
    for (emp_id, week), hours in hours_week.items():
        if hours > MAX_HOURS_PER_WEEK:
            soft_violations['over_30h'].append({
                'employee': emp[emp_id].name,
                'week': week + 1, 'hours': hours,
                'overtime': hours - MAX_HOURS_PER_WEEK
            })
    
    # Kiểm tra ràng buộc mềm: thời gian nghỉ
    for emp_id, seq in timeline.items():
        seq.sort(key=lambda x: x[0]*24 + x[1].start)
        for j in range(1, len(seq)):
            prev_d, prev_s = seq[j-1]
            cur_d, cur_s = seq[j]
            rest = (cur_d*24 + cur_s.start) - (prev_d*24 + prev_s.end)
            if rest < MIN_REST_HOURS:
                soft_violations['no_rest_12h'].append({
                    'employee': emp[emp_id].name,
                    'from': f"Ngày {prev_d+1} ca {prev_s.name}",
                    'to': f"Ngày {cur_d+1} ca {cur_s.name}",
                    'rest_hours': rest, 'missing': MIN_REST_HOURS - rest
                })
    
    # Kiểm tra ràng buộc mềm: giờ làm/tháng
    for emp_id, hours in soft_stats['total_hours'].items():
        if hours > MAX_HOURS_PER_MONTH:
            soft_violations['over_monthly'].append({
                'employee': emp[emp_id].name,
                'hours': hours,
                'overtime': hours - MAX_HOURS_PER_MONTH
            })
        if hours < MIN_HOURS_PER_MONTH:
            soft_violations['under_monthly'].append({
                'employee': emp[emp_id].name,
                'hours': hours,
                'shortage': MIN_HOURS_PER_MONTH - hours
            })
    
    total_hours_list = list(soft_stats['total_hours'].values())
    shift_counts_list = list(soft_stats['shift_counts'].values())
    
    soft_metrics = {
        'avg_hours': np.mean(total_hours_list) if total_hours_list else 0,
        'std_hours': np.std(total_hours_list) if total_hours_list else 0,
        'min_hours': min(total_hours_list) if total_hours_list else 0,
        'max_hours': max(total_hours_list) if total_hours_list else 0,
        'avg_shifts': np.mean(shift_counts_list) if shift_counts_list else 0,
        'std_shifts': np.std(shift_counts_list) if shift_counts_list else 0,
        'fairness_score': sum(abs(h - np.mean(total_hours_list)) for h in total_hours_list) if total_hours_list else 0
    }
    
    return hard_violations, soft_violations, soft_metrics, soft_stats


def print_constraint_report(hard_violations, soft_violations, soft_metrics, soft_stats):
    print("\n" + "="*80)
    print("BÁO CÁO KIỂM TRA RÀNG BUỘC")
    print("="*80)
    
    print("\nRÀNG BUỘC CỨNG (HARD CONSTRAINTS - PHẢI ĐẢM BẢO)")
    print("-"*80)
    
    total_hard = sum(len(v) for v in hard_violations.values())
    
    if total_hard == 0:
        print("Không có vi phạm ràng buộc cứng nào!")
    else:
        print(f"Tổng số vi phạm CỨNG: {total_hard}\n")
        
        if hard_violations['no_doctor']:
            print(f"Thiếu bác sĩ: {len(hard_violations['no_doctor'])} ca")
            for v in hard_violations['no_doctor'][:3]:
                print(f"Ngày {v['day']}, ca {v['shift']}, {v['room']} ({v['dept']}): "
                      f"Cần {v['required']}, chỉ có {v['actual']}")
            if len(hard_violations['no_doctor']) > 3:
                print(f"   ... và {len(hard_violations['no_doctor']) - 3} ca khác\n")
        
        if hard_violations['wrong_dept']:
            print(f"Phân công sai khoa: {len(hard_violations['wrong_dept'])} lượt")
            for v in hard_violations['wrong_dept'][:3]:
                print(f"   • {v['employee']}: được phân vào {v['assigned_dept']} "
                      f"(khoa đúng: {v['correct_dept']})")
            if len(hard_violations['wrong_dept']) > 3:
                print(f"   ... và {len(hard_violations['wrong_dept']) - 3} lượt khác\n")
    
    print("\nRÀNG BUỘC MỀM (SOFT CONSTRAINTS)")
    print("-"*80)
    
    total_soft = sum(len(v) for v in soft_violations.values())
    print(f"Tổng số vi phạm MỀM: {total_soft}")
    
    if soft_violations['over_30h']:
        print(f"\nVượt 30h/tuần: {len(soft_violations['over_30h'])} lượt")
        total_overtime = sum(v['overtime'] for v in soft_violations['over_30h'])
        print(f"   Tổng giờ vượt: {total_overtime}h")
        for v in soft_violations['over_30h'][:3]:
            print(f"{v['employee']}: Tuần {v['week']} = {v['hours']}h (+{v['overtime']}h)")
    
    if soft_violations['no_rest_12h']:
        print(f"\nThiếu nghỉ 12h giữa ca: {len(soft_violations['no_rest_12h'])} lượt")
        for v in soft_violations['no_rest_12h'][:3]:
            print(f"{v['employee']}: {v['from']} → {v['to']} = {v['rest_hours']}h nghỉ")
    
    if soft_violations['over_monthly']:
        print(f"\nVượt 120h/tháng: {len(soft_violations['over_monthly'])} người")
        for v in soft_violations['over_monthly'][:3]:
            print(f"{v['employee']}: {v['hours']}h (+{v['overtime']}h)")
    
    if soft_violations['under_monthly']:
        print(f"\nDưới 60h/tháng: {len(soft_violations['under_monthly'])} người")
        for v in soft_violations['under_monthly'][:3]:
            print(f"{v['employee']}: {v['hours']}h (-{v['shortage']}h)")
    
    print(f"\nThống kê công bằng:")
    print(f"   Giờ làm việc trung bình: {soft_metrics['avg_hours']:.1f}h (σ = {soft_metrics['std_hours']:.1f}h)")
    print(f"   Khoảng giờ làm việc: {soft_metrics['min_hours']:.0f}h - {soft_metrics['max_hours']:.0f}h")
    print(f"   Số ca trực trung bình: {soft_metrics['avg_shifts']:.1f} ca (σ = {soft_metrics['std_shifts']:.1f} ca)")
    print("="*80 + "\n")


def fitness(schedule, employees, dept_to_rooms, shifts, days, log=False):
    emp = {e.id: e for e in employees}
    hours_week = defaultdict(int)
    timeline = defaultdict(list)
    hard = defaultdict(int)
    soft = defaultdict(int)
    
    all_rooms = [room for rooms in dept_to_rooms.values() for room in rooms]
    
    for d in days:
        for s in shifts:
            for room in all_rooms:
                dept = get_room_department(room, dept_to_rooms)
                ids = schedule[d][s.name][room]
                
                doctors = [i for i in ids if emp[i].role == "doctor"]
                nurses  = [i for i in ids if emp[i].role == "nurse"]
                
                if len(doctors) < MIN_DOCTOR_PER_SHIFT:
                    hard["no_doctor"] += (MIN_DOCTOR_PER_SHIFT - len(doctors))
                if len(nurses) < MIN_NURSE_PER_SHIFT:
                    hard["no_nurse"] += (MIN_NURSE_PER_SHIFT - len(nurses))
                
                total_staff = len(doctors) + len(nurses)
                if total_staff < MIN_TOTAL_PER_SHIFT:
                    hard["less_than_5"] += (MIN_TOTAL_PER_SHIFT - total_staff)
                
                assigned_emps = [emp[i] for i in ids]
                has_senior = any(e.years_exp >= MIN_EXPERIENCE_YEARS for e in assigned_emps)
                if not has_senior:
                    hard["no_senior"] += 1
                
                for i in ids:
                    if emp[i].department != dept:
                        hard["wrong_dept"] += 1
                    if d in emp[i].days_off:
                        hard["day_off"] += 1
                    
                    week = d // 7
                    hours_week[(i, week)] += s.hours
                    timeline[i].append((d, s))
    
    # Tính soft constraint: over 30h/week
    for (i, _), h in hours_week.items():
        if h > MAX_HOURS_PER_WEEK:
            soft["over_30h"] += (h - MAX_HOURS_PER_WEEK)
    
    # Tính soft constraint: no rest 12h
    for i, seq in timeline.items():
        seq.sort(key=lambda x: x[0]*24 + x[1].start)
        for j in range(1, len(seq)):
            prev_d, prev_s = seq[j-1]
            cur_d, cur_s = seq[j]
            rest = (cur_d*24 + cur_s.start) - (prev_d*24 + prev_s.end)
            if rest < MIN_REST_HOURS:
                soft["no_rest_12h"] += 1
    
    # Tính soft constraint: over/under monthly hours
    total_hours = defaultdict(int)
    for (i, _), h in hours_week.items():
        total_hours[i] += h
    
    for i in total_hours.keys():
        if total_hours[i] > MAX_HOURS_PER_MONTH:
            soft["over_monthly"] += (total_hours[i] - MAX_HOURS_PER_MONTH)
        if total_hours[i] < MIN_HOURS_PER_MONTH:
            soft["under_monthly"] += (MIN_HOURS_PER_MONTH - total_hours[i])
    
    # Tính fairness penalty
    avg = np.mean(list(total_hours.values())) if total_hours else 0
    fairness = 0
    for h in total_hours.values():
        fairness += abs(h - avg)
    
    # Tổng điểm phạt
    total = (
        # Hard constraints - Phạt cực nặng
        hard["no_doctor"]     * W_NO_DOCTOR +
        hard["no_nurse"]      * W_NO_NURSE +
        hard["less_than_5"]   * W_LESS_5 +
        hard["no_senior"]     * W_NO_SENIOR +
        hard["wrong_dept"]    * W_WRONG_DEPT +
        hard["day_off"]       * W_DAY_OFF +
        # Soft constraints - Phạt nhẹ
        soft["over_30h"]      * W_OVER_30H +
        soft["no_rest_12h"]   * W_NO_REST +
        soft["over_monthly"]  * W_OVER_MONTHLY +
        soft["under_monthly"] * W_UNDER_MONTHLY +
        fairness * W_FAIRNESS
    )
    
    if log:
        return total, dict(hard), dict(soft), fairness
    return total


def tournament_selection(scored):
    pool_size = int(len(scored) * PARENT_POOL_RATIO)
    pool = scored[:pool_size]
    contenders = random.sample(pool, TOURNAMENT_K)
    contenders.sort(key=lambda x: x[0])
    return contenders[0][1]


def _create_valid_assignment(employees, dept, day):
    """Tạo assignment hợp lệ cho 1 khoa"""
    pool = [e for e in employees if e.department == dept and day not in e.days_off]
    doctors = [e for e in pool if e.role == "doctor"]
    nurses = [e for e in pool if e.role == "nurse"]
    
    selected_doctors = []
    selected_nurses = []
    
    if len(doctors) >= MIN_DOCTOR_PER_SHIFT:
        selected_doctors = random.sample(doctors, MIN_DOCTOR_PER_SHIFT)
    elif doctors:
        selected_doctors = doctors
    
    if len(nurses) >= MIN_NURSE_PER_SHIFT:
        selected_nurses = random.sample(nurses, MIN_NURSE_PER_SHIFT)
    elif nurses:
        selected_nurses = nurses
    
    assignment = [e.id for e in selected_doctors + selected_nurses]
    
    selected_all = selected_doctors + selected_nurses
    has_senior = any(e.years_exp >= MIN_EXPERIENCE_YEARS for e in selected_all)
    
    if not has_senior:
        seniors = [e for e in pool if e.years_exp >= MIN_EXPERIENCE_YEARS]
        if seniors:
            senior = random.choice(seniors)
            if senior.id not in assignment:
                assignment.append(senior.id)
    
    return assignment


def crossover_uniform(a, b, employees, dept_to_rooms):
    c = copy.deepcopy(a)
    emp = {e.id: e for e in employees}
    
    for d in a.keys():
        for s in a[d].keys():
            for room in a[d][s].keys():
                dept = get_room_department(room, dept_to_rooms)
                if random.random() < 0.5:
                    b_assignment = b[d][s][room]
                    
                    if len(b_assignment) >= MIN_TOTAL_PER_SHIFT:
                        doctors = [i for i in b_assignment if emp[i].role == "doctor"]
                        nurses = [i for i in b_assignment if emp[i].role == "nurse"]
                        has_senior = any(emp[i].years_exp >= MIN_EXPERIENCE_YEARS for i in b_assignment)
                        
                        if (len(doctors) >= MIN_DOCTOR_PER_SHIFT and 
                            len(nurses) >= MIN_NURSE_PER_SHIFT and 
                            has_senior):
                            c[d][s][room] = copy.deepcopy(b_assignment)
                        else:
                            c[d][s][room] = _create_valid_assignment(employees, dept, d)
                    else:
                        c[d][s][room] = _create_valid_assignment(employees, dept, d)
    
    return c


def mutate_scramble(ind, employees, dept_to_rooms, shifts, days, rate=0.3):
    if random.random() > rate:
        return ind
    
    d = random.choice(days)
    dept = random.choice(list(dept_to_rooms.keys()))
    rooms = dept_to_rooms[dept]
    
    if not rooms:
        return ind
    
    room = random.choice(rooms)
    
    assignments = []
    shift_names = []
    
    for s in shifts:
        assignment = ind[d][s.name].get(room, [])
        shift_names.append(s.name)
        
        if len(assignment) >= MIN_TOTAL_PER_SHIFT:
            assignments.append(assignment)
        else:
            assignments.append(_create_valid_assignment(employees, dept, d))
    
    if len(assignments) > 1:
        random.shuffle(assignments)
        
        for idx, s_name in enumerate(shift_names):
            if idx < len(assignments):
                ind[d][s_name][room] = assignments[idx]
    
    return ind


def mutate_balance_hours(ind, employees, dept_to_rooms, shifts, days, rate=0.3):
    if random.random() > rate:
        return ind
    
    emp = {e.id: e for e in employees}
    
    emp_hours = defaultdict(int)
    for d in days:
        for s in shifts:
            for room in [r for rooms in dept_to_rooms.values() for r in rooms]:
                for emp_id in ind[d][s.name][room]:
                    emp_hours[emp_id] += s.hours
    
    if not emp_hours:
        return ind
    
    avg_hours = np.mean(list(emp_hours.values()))
    overworked = [(emp_id, h) for emp_id, h in emp_hours.items() if h > avg_hours + 10]
    underworked = [(emp_id, h) for emp_id, h in emp_hours.items() if h < avg_hours - 10]
    
    if not overworked or not underworked:
        return ind
    
    over_emp_id, _ = random.choice(overworked)
    under_emp_id, _ = random.choice(underworked)
    
    over_emp = emp[over_emp_id]
    under_emp = emp[under_emp_id]
    
    if over_emp.role != under_emp.role or over_emp.department != under_emp.department:
        return ind
    
    for d in days:
        if d in under_emp.days_off:
            continue
        for s in shifts:
            for room in dept_to_rooms.get(over_emp.department, []):
                assignment = ind[d][s.name][room]
                if over_emp_id in assignment and under_emp_id not in assignment:
                    idx = assignment.index(over_emp_id)
                    assignment[idx] = under_emp_id
                    return ind
    
    return ind


def hill_climb(ind, employees, dept_to_rooms, shifts, days, steps=50):
    best = copy.deepcopy(ind)
    best_fit = fitness(best, employees, dept_to_rooms, shifts, days)
    
    all_rooms = [room for rooms in dept_to_rooms.values() for room in rooms]
    
    for _ in range(steps):
        neigh = copy.deepcopy(best)
        
        d = random.choice(days)
        room = random.choice(all_rooms)
        dept = get_room_department(room, dept_to_rooms)
        
        if len(shifts) >= 2:
            s1, s2 = random.sample(shifts, 2)
            
            assign1 = neigh[d][s1.name].get(room, [])
            assign2 = neigh[d][s2.name].get(room, [])
            
            if len(assign1) < MIN_TOTAL_PER_SHIFT:
                assign1 = _create_valid_assignment(employees, dept, d)
            
            if len(assign2) < MIN_TOTAL_PER_SHIFT:
                assign2 = _create_valid_assignment(employees, dept, d)
            
            neigh[d][s1.name][room] = assign2
            neigh[d][s2.name][room] = assign1
        
        f = fitness(neigh, employees, dept_to_rooms, shifts, days)
        if f < best_fit:
            best, best_fit = neigh, f
    
    return best


def export_calendar_to_excel(schedule, employees, dept_to_rooms, shifts, days, filename="lich_truc.xlsx"):
    """Xuất lịch trực theo khoa và phòng"""
    emp_dict = {e.id: e for e in employees}
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        for dept, rooms in dept_to_rooms.items():
            for room in rooms:
                data = []
                
                for day in days:
                    row = {"Ngày": f"Ngày {day + 1}"}
                    
                    for shift in shifts:
                        assignment = schedule[day][shift.name].get(room, [])
                        
                        doctors = [emp_dict[i] for i in assignment if emp_dict[i].role == "doctor"]
                        nurses = [emp_dict[i] for i in assignment if emp_dict[i].role == "nurse"]
                        
                        doctor_names = "\n".join([f"{d.name} ({d.years_exp}y)" for d in doctors]) if doctors else "THIẾU"
                        nurse_names = "\n".join([f"{n.name} ({n.years_exp}y)" for n in nurses]) if nurses else "THIẾU"
                        
                        row[f"{shift.name}\nBác sĩ"] = doctor_names
                        row[f"{shift.name}\nĐiều dưỡng"] = nurse_names
                    
                    data.append(row)
                
                df = pd.DataFrame(data)
                
                sheet_name = f"{dept[:10]}_{room[:15]}".replace("/", "-")[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                worksheet = writer.sheets[sheet_name]
                
                worksheet.column_dimensions['A'].width = 12
                for col in range(2, len(df.columns) + 2):
                    worksheet.column_dimensions[chr(64 + col)].width = 20
                
                from openpyxl.styles import Alignment, PatternFill, Font
                
                for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1):
                    for cell in row:
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                
                for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df) + 1), start=2):
                    fill_color = "E7E6E6" if idx % 2 == 0 else "FFFFFF"
                    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    for cell in row:
                        cell.fill = fill
    
    print(f"\nĐã xuất lịch trực ra file: {filename}")


def export_employee_hours_to_excel(schedule, employees, dept_to_rooms, shifts, days, filename="gio_lam_nhan_vien.xlsx"):
    """Xuất thời gian làm việc của từng nhân viên"""
    emp_dict = {e.id: e for e in employees}
    
    employee_hours = defaultdict(lambda: {
        'total_hours': 0,
        'total_shifts': 0,
        'weeks': defaultdict(int),
        'shifts_detail': []
    })
    
    all_rooms = [room for rooms in dept_to_rooms.values() for room in rooms]
    
    for d in days:
        for s in shifts:
            for room in all_rooms:
                assignment = schedule[d][s.name].get(room, [])
                for emp_id in assignment:
                    employee_hours[emp_id]['total_hours'] += s.hours
                    employee_hours[emp_id]['total_shifts'] += 1
                    week = d // 7
                    employee_hours[emp_id]['weeks'][week] += s.hours
                    employee_hours[emp_id]['shifts_detail'].append({
                        'day': d + 1,
                        'shift': s.name,
                        'room': room,
                        'hours': s.hours
                    })
    
    summary_data = []
    for emp_id, stats in employee_hours.items():
        emp = emp_dict[emp_id]
        row = {
            'ID': emp.id,
            'Tên': emp.name,
            'Chức vụ': 'Bác sĩ' if emp.role == 'doctor' else 'Điều dưỡng',
            'Khoa': emp.department,
            'Kinh nghiệm (năm)': emp.years_exp,
            'Tổng giờ': stats['total_hours'],
            'Số ca trực': stats['total_shifts'],
            'TB giờ/tuần': round(stats['total_hours'] / len(stats['weeks']), 1) if stats['weeks'] else 0
        }
        
        for week in range(max(days) // 7 + 1):
            row[f'Tuần {week + 1}'] = stats['weeks'].get(week, 0)
        
        summary_data.append(row)
    
    df_summary = pd.DataFrame(summary_data)
    df_summary = df_summary.sort_values(['Khoa', 'Chức vụ', 'Tổng giờ'], ascending=[True, True, False])
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df_summary.to_excel(writer, sheet_name='Tổng hợp', index=False)
        
        worksheet = writer.sheets['Tổng hợp']
        
        from openpyxl.styles import Alignment, PatternFill, Font
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        for idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df_summary) + 1), start=2):
            fill_color = "E7E6E6" if idx % 2 == 0 else "FFFFFF"
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for emp_id, stats in employee_hours.items():
            emp = emp_dict[emp_id]
            sheet_name = f"{emp.name}"[:31]
            
            detail_data = []
            for shift_info in stats['shifts_detail']:
                detail_data.append({
                    'Ngày': shift_info['day'],
                    'Ca': shift_info['shift'],
                    'Phòng': shift_info['room'],
                    'Giờ làm': shift_info['hours']
                })
            
            df_detail = pd.DataFrame(detail_data)
            df_detail.to_excel(writer, sheet_name=sheet_name, index=False)
            
            ws_detail = writer.sheets[sheet_name]
            
            for cell in ws_detail[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            for col in range(1, 5):
                ws_detail.column_dimensions[chr(64 + col)].width = 15
            
            for idx, row in enumerate(ws_detail.iter_rows(min_row=2, max_row=len(df_detail) + 1), start=2):
                fill_color = "E7E6E6" if idx % 2 == 0 else "FFFFFF"
                fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                for cell in row:
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
    
    print(f"\nĐã xuất thời gian làm việc nhân viên ra file: {filename}")


def export_violations_to_excel(hard_violations, soft_violations, soft_metrics, filename="bao_cao_vi_pham.xlsx"):
    """Xuất báo cáo vi phạm ra Excel"""
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        from openpyxl.styles import Alignment, PatternFill, Font
        
        # Summary sheet
        summary_data = []
        summary_data.append({'Loại': 'HARD CONSTRAINTS', 'Vi phạm': '', 'Số lượng': ''})
        for vtype, violations in hard_violations.items():
            summary_data.append({
                'Loại': '',
                'Vi phạm': vtype,
                'Số lượng': len(violations)
            })
        
        summary_data.append({'Loại': '', 'Vi phạm': '', 'Số lượng': ''})
        summary_data.append({'Loại': 'SOFT CONSTRAINTS', 'Vi phạm': '', 'Số lượng': ''})
        for vtype, violations in soft_violations.items():
            summary_data.append({
                'Loại': '',
                'Vi phạm': vtype,
                'Số lượng': len(violations)
            })
        
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name='Tổng hợp', index=False)
        
        ws = writer.sheets['Tổng hợp']
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Hard violations sheets
        for vtype, violations in hard_violations.items():
            if violations:
                df = pd.DataFrame(violations)
                sheet_name = f"HARD_{vtype}"[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                ws = writer.sheets[sheet_name]
                for cell in ws[1]:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        # Soft violations sheets
        for vtype, violations in soft_violations.items():
            if violations:
                df = pd.DataFrame(violations)
                sheet_name = f"SOFT_{vtype}"[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                ws = writer.sheets[sheet_name]
                for cell in ws[1]:
                    cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
                    cell.font = Font(bold=True, color="FFFFFF")
        
        # Soft metrics sheet
        soft_data = [{
            'Metric': k,
            'Value': v
        } for k, v in soft_metrics.items()]
        
        df_soft = pd.DataFrame(soft_data)
        df_soft.to_excel(writer, sheet_name='Thống kê', index=False)
    
    print(f"\nĐã xuất báo cáo vi phạm ra file: {filename}")


def print_calendar_console(schedule, employees, dept_to_rooms, shifts, days, dept_name):
    """In lịch trực của 1 khoa ra console"""
    emp_dict = {e.id: e for e in employees}
    rooms = dept_to_rooms.get(dept_name, [])
    
    if not rooms:
        print(f"Không tìm thấy khoa: {dept_name}")
        return
    
    print(f"\n{'='*100}")
    print(f"LỊCH TRỰC - {dept_name}".center(100))
    print(f"{'='*100}")
    
    for room in rooms:
        print(f"\n {room}")
        print("-" * 100)
        
        header = f"{'Ngày':<10}"
        for shift in shifts:
            header += f"| {shift.name:^28}"
        print(header)
        print("-" * 100)
        
        for day in days[:7]:
            row_lines = []
            
            for shift in shifts:
                assignment = schedule[day][shift.name].get(room, [])
                
                doctors = [emp_dict[i] for i in assignment if emp_dict[i].role == "doctor"]
                nurses = [emp_dict[i] for i in assignment if emp_dict[i].role == "nurse"]
                
                cell_text = []
                if doctors:
                    cell_text.append(f"BS: {', '.join([f'{d.name}({d.years_exp}y)' for d in doctors])}")
                if nurses:
                    cell_text.append(f"DD: {', '.join([f'{n.name}({n.years_exp}y)' for n in nurses[:2]])}")
                
                row_lines.append("\n".join(cell_text) if cell_text else "THIẾU")
            
            print(f"Ngày {day+1:<5}", end="")
            for cell in row_lines:
                print(f"| {cell:^28}", end="")
            print()
    
    print("=" * 100)


def main():
    employees, dept_to_rooms, shifts, days = generate_sample_data()
    
    population = [create_individual(employees, dept_to_rooms, shifts, days)
                  for _ in range(POPULATION_SIZE)]
    
    best_fit = float("inf")
    stagnation = 0
    history = []
    
    for gen in range(GENERATIONS):
        scored = [(fitness(ind, employees, dept_to_rooms, shifts, days), ind)
                  for ind in population]
        scored.sort(key=lambda x: x[0])
        
        best = scored[0][1]
        fit, hard, soft, fairness = fitness(best, employees, dept_to_rooms, shifts, days, log=True)
        
        print(f"Gen {gen:3d} | Best={fit:.0f} | HARD={hard} | SOFT={soft}")
        history.append(fit)
        
        if fit < best_fit:
            best_fit = fit
            stagnation = 0
        else:
            stagnation += 1
        
        if stagnation >= STAGNATION_LIMIT:
            print("  ↳ Hill Climbing triggered")
            best = hill_climb(best, employees, dept_to_rooms, shifts, days, HILL_CLIMB_STEPS)
            stagnation = 0
        
        new_pop = [copy.deepcopy(scored[i][1]) for i in range(ELITE_SIZE)]
        
        while len(new_pop) < POPULATION_SIZE:
            p1 = tournament_selection(scored)
            p2 = tournament_selection(scored)
            child = crossover_uniform(p1, p2, employees, dept_to_rooms)
            child = mutate_scramble(child, employees, dept_to_rooms, shifts, days, MUTATION_RATE)
            child = mutate_balance_hours(child, employees, dept_to_rooms, shifts, days, 0.3)
            new_pop.append(child)
        
        population = new_pop
    
    # Vẽ đồ thị convergence
    plt.figure(figsize=(10, 6))
    plt.plot(history)
    plt.xlabel("Generation")
    plt.ylabel("Best Fitness")
    plt.title("GA Convergence - Soft Constraints Enhanced")
    plt.grid()
    plt.tight_layout()
    plt.savefig("ga_convergence.png", dpi=150)
    plt.show()
    
    # Lấy solution tốt nhất
    best_schedule = scored[0][1]
    
    # Kiểm tra ràng buộc chi tiết
    hard_violations, soft_violations, soft_metrics, soft_stats = check_constraints_detailed(
        best_schedule, employees, dept_to_rooms, shifts, days
    )
    
    # In báo cáo
    print_constraint_report(hard_violations, soft_violations, soft_metrics, soft_stats)
    
    # Export các file
    export_calendar_to_excel(best_schedule, employees, dept_to_rooms, shifts, days, "lich_truc_benh_vien.xlsx")
    export_employee_hours_to_excel(best_schedule, employees, dept_to_rooms, shifts, days, "gio_lam_nhan_vien.xlsx")
    export_violations_to_excel(hard_violations, soft_violations, soft_metrics, "bao_cao_vi_pham.xlsx")
    
    # In lịch console cho khoa đầu tiên
    first_dept = list(dept_to_rooms.keys())[0]
    print_calendar_console(best_schedule, employees, dept_to_rooms, shifts, days, first_dept)
    
    print("\nHoàn tất! Đã tạo 3 file:")
    print("   lich_truc_benh_vien.xlsx")
    print("   gio_lam_nhan_vien.xlsx")
    print("   bao_cao_vi_pham.xlsx")


if __name__ == "__main__":
    main()