"""
Ứng dụng GUI tkinter cho hệ thống lập lịch bệnh viện
Sử dụng thuật toán GA từ schedule-v7.py
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import sys
import os
import importlib.util
import threading
import queue
import copy
from collections import defaultdict

# Import matplotlib cho biểu đồ
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure

# Import openpyxl cho Excel export
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import từ file schedule-v7.py
import importlib.util
spec = importlib.util.spec_from_file_location("ga_module", 
                                               os.path.join(os.path.dirname(__file__), "schedule-v7.py"))
ga_module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(ga_module)


class HospitalScheduleApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Hệ thống Lập lịch Trực Bệnh viện")
        self.root.geometry("1200x800")
        
        # Load cấu hình từ module gốc
        self.config = self.load_config_from_module()
        
        # Data
        self.employees = []
        self.dept_to_rooms = {}
        self.shifts = []
        self.days = []
        self.best_schedule = None
        
        # GA running state
        self.is_running = False
        self.history = []
        self.output_queue = queue.Queue()
        
        # Setup UI
        self.setup_main_ui()
        
        # Start checking queue
        self.root.after(100, self.check_queue)
        
    def load_config_from_module(self):
        """Load cấu hình từ module schedule_v7"""
        return {
            # Thông tin chung
            'NUM_DAYS': ga_module.NUM_DAYS,
            'DEPARTMENTS': list(ga_module.DEPARTMENTS.keys()),
            'DOCTORS_PER_DEPARTMENT': ga_module.DOCTORS_PER_DEPARTMENT,
            'NURSES_PER_DEPARTMENT': ga_module.NURSES_PER_DEPARTMENT,
            'SENIOR_DOCTOR_RATIO': ga_module.SENIOR_DOCTOR_RATIO,
            'SENIOR_NURSE_RATIO': ga_module.SENIOR_NURSE_RATIO,
            
            # Hard constraints
            'MIN_DOCTOR_PER_SHIFT': ga_module.MIN_DOCTOR_PER_SHIFT,
            'MIN_NURSE_PER_SHIFT': ga_module.MIN_NURSE_PER_SHIFT,
            'MIN_TOTAL_PER_SHIFT': ga_module.MIN_TOTAL_PER_SHIFT,
            'MIN_EXPERIENCE_YEARS': ga_module.MIN_EXPERIENCE_YEARS,
            
            # Soft constraints
            'MAX_HOURS_PER_WEEK': ga_module.MAX_HOURS_PER_WEEK,
            'MIN_REST_HOURS': ga_module.MIN_REST_HOURS,
            'MAX_HOURS_PER_MONTH': ga_module.MAX_HOURS_PER_MONTH,
            'MIN_HOURS_PER_MONTH': ga_module.MIN_HOURS_PER_MONTH,
            
            # GA parameters
            'POPULATION_SIZE': ga_module.POPULATION_SIZE,
            'GENERATIONS': ga_module.GENERATIONS,
            'ELITE_SIZE': ga_module.ELITE_SIZE,
            'TOURNAMENT_K': ga_module.TOURNAMENT_K,
            'PARENT_POOL_RATIO': ga_module.PARENT_POOL_RATIO,
            'MUTATION_RATE': ga_module.MUTATION_RATE,
            'STAGNATION_LIMIT': ga_module.STAGNATION_LIMIT,
            'HILL_CLIMB_STEPS': ga_module.HILL_CLIMB_STEPS,
            
            # Penalty weights
            'W_NO_DOCTOR': ga_module.W_NO_DOCTOR,
            'W_NO_NURSE': ga_module.W_NO_NURSE,
            'W_LESS_5': ga_module.W_LESS_5,
            'W_NO_SENIOR': ga_module.W_NO_SENIOR,
            'W_WRONG_DEPT': ga_module.W_WRONG_DEPT,
            'W_DAY_OFF': ga_module.W_DAY_OFF,
            'W_OVER_30H': ga_module.W_OVER_30H,
            'W_NO_REST': ga_module.W_NO_REST,
            'W_OVER_MONTHLY': ga_module.W_OVER_MONTHLY,
            'W_UNDER_MONTHLY': ga_module.W_UNDER_MONTHLY,
            'W_FAIRNESS': ga_module.W_FAIRNESS,
        }
    
    def setup_main_ui(self):
        """Thiết lập giao diện chính"""
        # Header
        header_frame = ttk.Frame(self.root)
        header_frame.pack(fill="x", padx=10, pady=10)
        
        title_label = ttk.Label(header_frame, 
                               text="HỆ THỐNG LẬP LỊCH TRỰC BỆNH VIỆN",
                               font=('Arial', 18, 'bold'),
                               foreground='#2E86AB')
        title_label.pack()
        
        subtitle_label = ttk.Label(header_frame,
                                   text="Sử dụng thuật toán di truyền (Genetic Algorithm)",
                                   font=('Arial', 10),
                                   foreground='#666')
        subtitle_label.pack()
        
        # Notebook (tabs)
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=5)
        
        # Tạo các tab
        self.tab1 = ttk.Frame(self.notebook)
        self.tab2 = ttk.Frame(self.notebook)
        self.tab3 = ttk.Frame(self.notebook)
        self.tab4 = ttk.Frame(self.notebook)
        
        self.notebook.add(self.tab1, text='⚙️  Cấu hình')
        self.notebook.add(self.tab2, text='📅  Đăng ký nghỉ')
        self.notebook.add(self.tab3, text='▶️  Chạy & Theo dõi')
        self.notebook.add(self.tab4, text='📊  Dashboard')
        
        # Setup Tab 1
        self.setup_tab1_config()
        
        # Setup Tab 2
        self.setup_tab2_dayoff()
        
        # Setup Tab 3
        self.setup_tab3_run()
        
        # Setup Tab 4
        self.setup_tab4_dashboard()
    
    def setup_tab1_config(self):
        """Tab 1: Cấu hình tham số"""
        # Main container với scrollbar
        main_container = ttk.Frame(self.tab1)
        main_container.pack(fill="both", expand=True)
        
        # Canvas và Scrollbar
        canvas = tk.Canvas(main_container, bg='white')
        scrollbar = ttk.Scrollbar(main_container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Mouse wheel scroll
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # Pack
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Dictionary lưu các entry widgets
        self.config_entries = {}
        
        # ===== SECTION 1: Thông tin chung =====
        section1 = ttk.LabelFrame(scrollable_frame, text="📋 Thông tin chung", padding="15")
        section1.pack(fill="x", padx=15, pady=10)
        
        common_params = [
            ('NUM_DAYS', 'Số ngày lập lịch', 'int', 'Tổng số ngày cần lập lịch (1-365)'),
            ('DOCTORS_PER_DEPARTMENT', 'Số bác sĩ mỗi khoa', 'int', 'Số lượng bác sĩ trong mỗi khoa'),
            ('NURSES_PER_DEPARTMENT', 'Số điều dưỡng mỗi khoa', 'int', 'Số lượng điều dưỡng trong mỗi khoa'),
            ('SENIOR_DOCTOR_RATIO', 'Tỷ lệ bác sĩ có kinh nghiệm (0-1)', 'float', 'Tỷ lệ bác sĩ có ≥5 năm kinh nghiệm (0.0-1.0)'),
            ('SENIOR_NURSE_RATIO', 'Tỷ lệ điều dưỡng có kinh nghiệm (0-1)', 'float', 'Tỷ lệ điều dưỡng có ≥5 năm kinh nghiệm (0.0-1.0)'),
        ]
        
        for i, (key, label, vtype, tooltip) in enumerate(common_params):
            # Label
            lbl = ttk.Label(section1, text=label + ":", font=('Arial', 9))
            lbl.grid(row=i, column=0, sticky="w", pady=5, padx=(0, 10))
            
            # Entry
            entry = ttk.Entry(section1, width=15, font=('Arial', 9))
            entry.insert(0, str(self.config[key]))
            entry.grid(row=i, column=1, pady=5, padx=(0, 10))
            self.config_entries[key] = entry
            
            # Tooltip
            tooltip_label = ttk.Label(section1, text=f"ℹ️ {tooltip}", 
                                     font=('Arial', 8), foreground='gray')
            tooltip_label.grid(row=i, column=2, sticky="w", pady=5)
        
        # ===== SECTION 2: Ràng buộc cứng =====
        section2 = ttk.LabelFrame(scrollable_frame, text="🔴 Ràng buộc cứng (Hard Constraints)", padding="15")
        section2.pack(fill="x", padx=15, pady=10)
        
        ttk.Label(section2, text="Các ràng buộc này PHẢI được thỏa mãn, nếu vi phạm sẽ bị phạt rất nặng",
                 font=('Arial', 8, 'italic'), foreground='red').grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 10))
        
        hard_params = [
            ('MIN_DOCTOR_PER_SHIFT', 'Số bác sĩ tối thiểu mỗi ca', 'int', 'Mỗi ca phải có ít nhất bao nhiêu bác sĩ'),
            ('MIN_NURSE_PER_SHIFT', 'Số điều dưỡng tối thiểu mỗi ca', 'int', 'Mỗi ca phải có ít nhất bao nhiêu điều dưỡng'),
            ('MIN_TOTAL_PER_SHIFT', 'Tổng nhân viên tối thiểu mỗi ca', 'int', 'Tổng số nhân viên tối thiểu mỗi ca'),
            ('MIN_EXPERIENCE_YEARS', 'Số năm kinh nghiệm tối thiểu', 'int', 'Mỗi ca phải có ít nhất 1 người có kinh nghiệm này'),
        ]
        
        for i, (key, label, vtype, tooltip) in enumerate(hard_params):
            lbl = ttk.Label(section2, text=label + ":", font=('Arial', 9))
            lbl.grid(row=i+1, column=0, sticky="w", pady=5, padx=(0, 10))
            
            entry = ttk.Entry(section2, width=15, font=('Arial', 9))
            entry.insert(0, str(self.config[key]))
            entry.grid(row=i+1, column=1, pady=5, padx=(0, 10))
            self.config_entries[key] = entry
            
            tooltip_label = ttk.Label(section2, text=f"ℹ️ {tooltip}", 
                                     font=('Arial', 8), foreground='gray')
            tooltip_label.grid(row=i+1, column=2, sticky="w", pady=5)
        
        # ===== SECTION 3: Ràng buộc mềm =====
        section3 = ttk.LabelFrame(scrollable_frame, text="🟡 Ràng buộc mềm (Soft Constraints)", padding="15")
        section3.pack(fill="x", padx=15, pady=10)
        
        ttk.Label(section3, text="Các ràng buộc này NÊN được thỏa mãn, vi phạm sẽ bị phạt nhẹ hơn",
                 font=('Arial', 8, 'italic'), foreground='orange').grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 10))
        
        soft_params = [
            ('MAX_HOURS_PER_WEEK', 'Số giờ tối đa mỗi tuần', 'int', 'Nhân viên không nên làm quá số giờ này/tuần'),
            ('MIN_REST_HOURS', 'Số giờ nghỉ tối thiểu giữa ca', 'int', 'Khoảng cách tối thiểu giữa 2 ca liên tiếp'),
            ('MAX_HOURS_PER_MONTH', 'Số giờ tối đa mỗi tháng', 'int', 'Nhân viên không nên làm quá số giờ này/tháng'),
            ('MIN_HOURS_PER_MONTH', 'Số giờ tối thiểu mỗi tháng', 'int', 'Nhân viên nên làm ít nhất số giờ này/tháng'),
        ]
        
        for i, (key, label, vtype, tooltip) in enumerate(soft_params):
            lbl = ttk.Label(section3, text=label + ":", font=('Arial', 9))
            lbl.grid(row=i+1, column=0, sticky="w", pady=5, padx=(0, 10))
            
            entry = ttk.Entry(section3, width=15, font=('Arial', 9))
            entry.insert(0, str(self.config[key]))
            entry.grid(row=i+1, column=1, pady=5, padx=(0, 10))
            self.config_entries[key] = entry
            
            tooltip_label = ttk.Label(section3, text=f"ℹ️ {tooltip}", 
                                     font=('Arial', 8), foreground='gray')
            tooltip_label.grid(row=i+1, column=2, sticky="w", pady=5)
        
        # ===== SECTION 4: Tham số GA =====
        section4 = ttk.LabelFrame(scrollable_frame, text="🧬 Tham số thuật toán di truyền (GA)", padding="15")
        section4.pack(fill="x", padx=15, pady=10)
        
        ttk.Label(section4, text="Các tham số điều khiển quá trình tiến hóa của thuật toán",
                 font=('Arial', 8, 'italic'), foreground='blue').grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 10))
        
        ga_params = [
            ('POPULATION_SIZE', 'Kích thước quần thể', 'int', 'Số lượng cá thể trong mỗi thế hệ'),
            ('GENERATIONS', 'Số thế hệ', 'int', 'Số thế hệ tiến hóa tối đa'),
            ('ELITE_SIZE', 'Số cá thể ưu tú', 'int', 'Số cá thể tốt nhất được giữ lại mỗi thế hệ'),
            ('TOURNAMENT_K', 'Kích thước tournament', 'int', 'Số cá thể tham gia tournament selection'),
            ('PARENT_POOL_RATIO', 'Tỷ lệ pool cha mẹ (0-1)', 'float', 'Tỷ lệ quần thể được chọn làm pool cha mẹ'),
            ('MUTATION_RATE', 'Tỷ lệ đột biến (0-1)', 'float', 'Xác suất xảy ra đột biến'),
            ('STAGNATION_LIMIT', 'Giới hạn stagnation', 'int', 'Số thế hệ không cải thiện trước khi hill climbing'),
            ('HILL_CLIMB_STEPS', 'Số bước hill climbing', 'int', 'Số bước leo đồi khi bị stagnation'),
        ]
        
        for i, (key, label, vtype, tooltip) in enumerate(ga_params):
            lbl = ttk.Label(section4, text=label + ":", font=('Arial', 9))
            lbl.grid(row=i+1, column=0, sticky="w", pady=5, padx=(0, 10))
            
            entry = ttk.Entry(section4, width=15, font=('Arial', 9))
            entry.insert(0, str(self.config[key]))
            entry.grid(row=i+1, column=1, pady=5, padx=(0, 10))
            self.config_entries[key] = entry
            
            tooltip_label = ttk.Label(section4, text=f"ℹ️ {tooltip}", 
                                     font=('Arial', 8), foreground='gray')
            tooltip_label.grid(row=i+1, column=2, sticky="w", pady=5)
        
        # ===== SECTION 5: Trọng số phạt =====
        section5 = ttk.LabelFrame(scrollable_frame, text="⚖️ Trọng số phạt (Penalty Weights)", padding="15")
        section5.pack(fill="x", padx=15, pady=10)
        
        ttk.Label(section5, text="Mức độ phạt cho từng loại vi phạm (số càng lớn = phạt càng nặng)",
                 font=('Arial', 8, 'italic'), foreground='purple').grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 10))
        
        # Hard constraint weights
        ttk.Label(section5, text="Phạt ràng buộc cứng:", font=('Arial', 9, 'bold')).grid(row=1, column=0, columnspan=3, sticky="w", pady=(10, 5))
        
        hard_weights = [
            ('W_NO_DOCTOR', 'Thiếu bác sĩ', 'int'),
            ('W_NO_NURSE', 'Thiếu điều dưỡng', 'int'),
            ('W_LESS_5', 'Thiếu 5 người', 'int'),
            ('W_NO_SENIOR', 'Thiếu người có kinh nghiệm', 'int'),
            ('W_WRONG_DEPT', 'Phân công sai khoa', 'int'),
            ('W_DAY_OFF', 'Vi phạm ngày nghỉ', 'int'),
        ]
        
        row = 2
        for key, label, vtype in hard_weights:
            lbl = ttk.Label(section5, text=label + ":", font=('Arial', 9))
            lbl.grid(row=row, column=0, sticky="w", pady=3, padx=(20, 10))
            
            entry = ttk.Entry(section5, width=15, font=('Arial', 9))
            entry.insert(0, str(self.config[key]))
            entry.grid(row=row, column=1, pady=3, padx=(0, 10))
            self.config_entries[key] = entry
            
            row += 1
        
        # Soft constraint weights
        ttk.Label(section5, text="Phạt ràng buộc mềm:", font=('Arial', 9, 'bold')).grid(row=row, column=0, columnspan=3, sticky="w", pady=(10, 5))
        row += 1
        
        soft_weights = [
            ('W_OVER_30H', 'Vượt 30h/tuần', 'int'),
            ('W_NO_REST', 'Thiếu nghỉ giữa ca', 'int'),
            ('W_OVER_MONTHLY', 'Vượt giờ tháng', 'int'),
            ('W_UNDER_MONTHLY', 'Thiếu giờ tháng', 'int'),
            ('W_FAIRNESS', 'Không công bằng', 'int'),
        ]
        
        for key, label, vtype in soft_weights:
            lbl = ttk.Label(section5, text=label + ":", font=('Arial', 9))
            lbl.grid(row=row, column=0, sticky="w", pady=3, padx=(20, 10))
            
            entry = ttk.Entry(section5, width=15, font=('Arial', 9))
            entry.insert(0, str(self.config[key]))
            entry.grid(row=row, column=1, pady=3, padx=(0, 10))
            self.config_entries[key] = entry
            
            row += 1
        
        # ===== BUTTON PANEL =====
        button_frame = ttk.Frame(scrollable_frame)
        button_frame.pack(fill="x", padx=15, pady=20)
        
        # Style cho buttons
        style = ttk.Style()
        style.configure('Action.TButton', font=('Arial', 10, 'bold'))
        
        ttk.Button(button_frame, text="💾 Lưu cấu hình", 
                  command=self.save_config,
                  style='Action.TButton',
                  width=20).pack(side="left", padx=5)
        
        ttk.Button(button_frame, text="🔄 Khôi phục mặc định", 
                  command=self.restore_defaults,
                  style='Action.TButton',
                  width=20).pack(side="left", padx=5)
        
        ttk.Button(button_frame, text="📂 Tạo dữ liệu mẫu", 
                  command=self.generate_sample_data,
                  style='Action.TButton',
                  width=20).pack(side="left", padx=5)
        
        ttk.Button(button_frame, text="💾 Lưu cấu hình ra file", 
                  command=self.export_config,
                  width=20).pack(side="left", padx=5)
        
        ttk.Button(button_frame, text="📂 Load cấu hình từ file", 
                  command=self.import_config,
                  width=20).pack(side="left", padx=5)
    
    def save_config(self):
        """Lưu cấu hình từ UI"""
        try:
            for key, entry in self.config_entries.items():
                value = entry.get().strip()
                
                # Xác định kiểu dữ liệu
                if 'RATIO' in key or 'RATE' in key or key == 'PARENT_POOL_RATIO':
                    self.config[key] = float(value)
                else:
                    self.config[key] = int(value)
            
            messagebox.showinfo("✅ Thành công", 
                              "Đã lưu cấu hình!\n\n"
                              "Cấu hình sẽ được áp dụng khi chạy thuật toán.")
        except ValueError as e:
            messagebox.showerror("❌ Lỗi", 
                               f"Giá trị không hợp lệ!\n\n"
                               f"Vui lòng kiểm tra lại các trường số.\n\n"
                               f"Chi tiết: {str(e)}")
    
    def restore_defaults(self):
        """Khôi phục cấu hình mặc định"""
        if messagebox.askyesno("⚠️ Xác nhận", 
                              "Khôi phục về cấu hình mặc định?\n\n"
                              "Tất cả thay đổi chưa lưu sẽ bị mất."):
            self.config = self.load_config_from_module()
            
            # Update UI
            for key, entry in self.config_entries.items():
                entry.delete(0, tk.END)
                entry.insert(0, str(self.config[key]))
            
            messagebox.showinfo("✅ Thành công", 
                              "Đã khôi phục cấu hình mặc định!")
    
    def generate_sample_data(self):
        """Tạo dữ liệu mẫu"""
        try:
            # Gọi hàm từ module gốc
            self.employees, self.dept_to_rooms, self.shifts, self.days = \
                ga_module.generate_sample_data()
            
            # Update tab 2 - employee list and dropdown
            if hasattr(self, 'employee_tree'):
                self.update_employee_list()
            
            # Update tab 4 - department dropdown
            if hasattr(self, 'dashboard_dept_combo'):
                self.dashboard_dept_combo['values'] = ["Tất cả"] + self.config['DEPARTMENTS']
            
            messagebox.showinfo("✅ Thành công", 
                              f"Đã tạo dữ liệu mẫu!\n\n"
                              f"📊 Thống kê:\n"
                              f"   • Số khoa: {len(self.dept_to_rooms)}\n"
                              f"   • Tổng số bác sĩ: {sum(e.role == 'doctor' for e in self.employees)}\n"
                              f"   • Tổng số điều dưỡng: {sum(e.role == 'nurse' for e in self.employees)}\n"
                              f"   • Tổng nhân viên: {len(self.employees)}\n\n"
                              f"Bạn có thể sang tab 'Đăng ký nghỉ' để xem và chỉnh sửa.")
        except Exception as e:
            messagebox.showerror("❌ Lỗi", 
                               f"Không thể tạo dữ liệu mẫu!\n\n"
                               f"Chi tiết: {str(e)}")
    
    def export_config(self):
        """Xuất cấu hình ra file"""
        filename = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            title="Lưu cấu hình"
        )
        
        if filename:
            try:
                with open(filename, 'w', encoding='utf-8') as f:
                    f.write("# Cấu hình hệ thống lập lịch bệnh viện\n")
                    f.write("# Tạo bởi Hospital Schedule App\n\n")
                    
                    for key, value in self.config.items():
                        f.write(f"{key} = {value}\n")
                
                messagebox.showinfo("✅ Thành công", 
                                  f"Đã lưu cấu hình vào:\n{filename}")
            except Exception as e:
                messagebox.showerror("❌ Lỗi", 
                                   f"Không thể lưu file!\n\n{str(e)}")
    
    def import_config(self):
        """Load cấu hình từ file"""
        filename = filedialog.askopenfilename(
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")],
            title="Load cấu hình"
        )
        
        if filename:
            try:
                with open(filename, 'r', encoding='utf-8') as f:
                    for line in f:
                        line = line.strip()
                        if line and not line.startswith('#'):
                            key, value = line.split('=')
                            key = key.strip()
                            value = value.strip()
                            
                            if key in self.config:
                                if 'RATIO' in key or 'RATE' in key:
                                    self.config[key] = float(value)
                                else:
                                    self.config[key] = int(value)
                                
                                # Update UI
                                if key in self.config_entries:
                                    self.config_entries[key].delete(0, tk.END)
                                    self.config_entries[key].insert(0, str(self.config[key]))
                
                messagebox.showinfo("✅ Thành công", 
                                  f"Đã load cấu hình từ:\n{filename}")
            except Exception as e:
                messagebox.showerror("❌ Lỗi", 
                                   f"Không thể load file!\n\n{str(e)}")
    
    def setup_tab2_dayoff(self):
        """Tab 2: Đăng ký nghỉ"""
        from datetime import datetime, timedelta
        
        # Main container
        main_container = ttk.PanedWindow(self.tab2, orient=tk.HORIZONTAL)
        main_container.pack(fill="both", expand=True, padx=10, pady=10)
        
        # ===== LEFT PANEL: Danh sách nhân viên =====
        left_panel = ttk.Frame(main_container)
        main_container.add(left_panel, weight=1)
        
        # Header
        header_frame = ttk.Frame(left_panel)
        header_frame.pack(fill="x", pady=(0, 10))
        
        ttk.Label(header_frame, text="👥 Danh sách nhân viên", 
                 font=('Arial', 12, 'bold')).pack(anchor="w")
        
        # Search box
        search_frame = ttk.Frame(left_panel)
        search_frame.pack(fill="x", pady=5)
        
        ttk.Label(search_frame, text="🔍 Tìm kiếm:", 
                 font=('Arial', 9)).pack(side="left", padx=(0, 5))
        
        self.search_var = tk.StringVar()
        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_var, 
                                      font=('Arial', 9))
        self.search_entry.pack(side="left", fill="x", expand=True, padx=(0, 5))
        self.search_var.trace('w', lambda *args: self.filter_employees())
        
        ttk.Button(search_frame, text="🔍", width=3,
                  command=self.filter_employees).pack(side="left")
        
        # Department filter
        filter_frame = ttk.Frame(left_panel)
        filter_frame.pack(fill="x", pady=5)
        
        ttk.Label(filter_frame, text="🏥 Khoa:", 
                 font=('Arial', 9)).pack(side="left", padx=(0, 5))
        
        self.dept_filter_var = tk.StringVar()
        self.dept_filter = ttk.Combobox(filter_frame, 
                                        textvariable=self.dept_filter_var,
                                        state="readonly", 
                                        font=('Arial', 9),
                                        width=25)
        self.dept_filter.pack(side="left", fill="x", expand=True)
        self.dept_filter.bind('<<ComboboxSelected>>', lambda e: self.filter_employees())
        
        # Role filter
        ttk.Label(filter_frame, text="👔 Chức vụ:", 
                 font=('Arial', 9)).pack(side="left", padx=(10, 5))
        
        self.role_filter_var = tk.StringVar(value="Tất cả")
        self.role_filter = ttk.Combobox(filter_frame,
                                        textvariable=self.role_filter_var,
                                        values=["Tất cả", "Bác sĩ", "Điều dưỡng"],
                                        state="readonly",
                                        font=('Arial', 9),
                                        width=15)
        self.role_filter.pack(side="left")
        self.role_filter.bind('<<ComboboxSelected>>', lambda e: self.filter_employees())
        
        # Employee count
        self.emp_count_label = ttk.Label(left_panel, text="Tổng: 0 nhân viên", 
                                        font=('Arial', 9), foreground='gray')
        self.emp_count_label.pack(fill="x", pady=5)
        
        # Employee Treeview
        tree_frame = ttk.Frame(left_panel)
        tree_frame.pack(fill="both", expand=True, pady=5)
        
        # Scrollbars
        tree_scroll_y = ttk.Scrollbar(tree_frame, orient="vertical")
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        self.employee_tree = ttk.Treeview(tree_frame,
                                         columns=("ID", "Tên", "Chức vụ", "Khoa", "Kinh nghiệm", "Ngày nghỉ"),
                                         show="headings",
                                         yscrollcommand=tree_scroll_y.set,
                                         xscrollcommand=tree_scroll_x.set,
                                         height=20)
        
        tree_scroll_y.config(command=self.employee_tree.yview)
        tree_scroll_x.config(command=self.employee_tree.xview)
        
        # Configure columns
        self.employee_tree.heading("ID", text="ID")
        self.employee_tree.heading("Tên", text="Tên")
        self.employee_tree.heading("Chức vụ", text="Chức vụ")
        self.employee_tree.heading("Khoa", text="Khoa")
        self.employee_tree.heading("Kinh nghiệm", text="Kinh nghiệm")
        self.employee_tree.heading("Ngày nghỉ", text="Số ngày nghỉ")
        
        self.employee_tree.column("ID", width=50, anchor="center")
        self.employee_tree.column("Tên", width=120, anchor="w")
        self.employee_tree.column("Chức vụ", width=100, anchor="center")
        self.employee_tree.column("Khoa", width=100, anchor="w")
        self.employee_tree.column("Kinh nghiệm", width=90, anchor="center")
        self.employee_tree.column("Ngày nghỉ", width=80, anchor="center")
        
        # Pack tree and scrollbars
        tree_scroll_y.pack(side="right", fill="y")
        tree_scroll_x.pack(side="bottom", fill="x")
        self.employee_tree.pack(side="left", fill="both", expand=True)
        
        # Bind selection event
        self.employee_tree.bind('<<TreeviewSelect>>', self.on_employee_select)
        
        # ===== RIGHT PANEL: Đăng ký nghỉ =====
        right_panel = ttk.Frame(main_container)
        main_container.add(right_panel, weight=1)
        
        # Header
        header_frame2 = ttk.Frame(right_panel)
        header_frame2.pack(fill="x", pady=(0, 10))
        
        ttk.Label(header_frame2, text="📅 Đăng ký ngày nghỉ", 
                 font=('Arial', 12, 'bold')).pack(anchor="w")
        
        # Selected employee info
        info_frame = ttk.LabelFrame(right_panel, text="👤 Nhân viên đã chọn", padding="10")
        info_frame.pack(fill="x", padx=5, pady=5)
        
        self.selected_emp_info = tk.StringVar(value="Chưa chọn nhân viên nào")
        ttk.Label(info_frame, textvariable=self.selected_emp_info,
                 font=('Arial', 10), foreground='blue').pack(anchor="w")
        
        self.selected_emp_id = None
        
        # Instructions
        instruction_frame = ttk.Frame(right_panel)
        instruction_frame.pack(fill="x", padx=5, pady=5)
        
        ttk.Label(instruction_frame, text="💡 Hướng dẫn: Nhấp vào các ô ngày để chọn/bỏ chọn ngày nghỉ",
                 font=('Arial', 9), foreground='green').pack(anchor="w")
        
        # Month navigation
        nav_frame = ttk.Frame(right_panel)
        nav_frame.pack(fill="x", padx=5, pady=10)
        
        ttk.Button(nav_frame, text="◀◀ Tháng trước",
                  command=lambda: self.change_month(-1)).pack(side="left", padx=5)
        
        self.current_date = datetime.now()
        self.month_label_var = tk.StringVar()
        ttk.Label(nav_frame, textvariable=self.month_label_var,
                 font=('Arial', 12, 'bold')).pack(side="left", expand=True)
        
        ttk.Button(nav_frame, text="Tháng sau ▶▶",
                  command=lambda: self.change_month(1)).pack(side="right", padx=5)
        
        ttk.Button(nav_frame, text="📅 Hôm nay",
                  command=self.go_to_today).pack(side="right", padx=5)
        
        # Calendar frame
        calendar_container = ttk.Frame(right_panel)
        calendar_container.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.calendar_frame = ttk.Frame(calendar_container)
        self.calendar_frame.pack(fill="both", expand=True)
        
        self.day_buttons = {}
        self.selected_days_set = set()
        
        # Button panel
        button_panel = ttk.Frame(right_panel)
        button_panel.pack(fill="x", padx=5, pady=10)
        
        ttk.Button(button_panel, text="💾 Lưu ngày nghỉ",
                  command=self.save_dayoffs,
                  style='Action.TButton').pack(side="left", padx=5)
        
        ttk.Button(button_panel, text="🗑️ Xóa tất cả",
                  command=self.clear_all_dayoffs,
                  style='Action.TButton').pack(side="left", padx=5)
        
        ttk.Button(button_panel, text="📊 Thống kê",
                  command=self.show_dayoff_stats,
                  style='Action.TButton').pack(side="left", padx=5)
        
        # Summary label
        self.dayoff_summary = tk.StringVar(value="Đã chọn: 0 ngày")
        ttk.Label(button_panel, textvariable=self.dayoff_summary,
                 font=('Arial', 10, 'bold'), foreground='red').pack(side="right", padx=10)
        
        # Initial setup
        self.update_employee_list()
        self.draw_calendar()
    
    def update_employee_list(self):
        """Cập nhật danh sách nhân viên"""
        # Clear tree
        for item in self.employee_tree.get_children():
            self.employee_tree.delete(item)
        
        if not self.employees:
            self.emp_count_label.config(text="⚠️ Chưa có dữ liệu. Vui lòng tạo dữ liệu mẫu ở tab Cấu hình!")
            self.dept_filter['values'] = []
            return
        
        # Update department filter
        depts = sorted(set(e.department for e in self.employees))
        self.dept_filter['values'] = ['Tất cả'] + depts
        if not self.dept_filter_var.get():
            self.dept_filter_var.set('Tất cả')
        
        # Add employees to tree
        for emp in self.employees:
            role_text = "Bác sĩ" if emp.role == "doctor" else "Điều dưỡng"
            self.employee_tree.insert("", "end", values=(
                emp.id,
                emp.name,
                role_text,
                emp.department,
                f"{emp.years_exp} năm",
                len(emp.days_off)
            ))
        
        self.emp_count_label.config(text=f"Tổng: {len(self.employees)} nhân viên")
    
    def filter_employees(self):
        """Lọc nhân viên theo điều kiện"""
        # Clear tree
        for item in self.employee_tree.get_children():
            self.employee_tree.delete(item)
        
        if not self.employees:
            return
        
        search_text = self.search_var.get().lower()
        dept_filter = self.dept_filter_var.get()
        role_filter = self.role_filter_var.get()
        
        count = 0
        for emp in self.employees:
            # Filter by search text
            if search_text and search_text not in emp.name.lower() and search_text not in str(emp.id):
                continue
            
            # Filter by department
            if dept_filter and dept_filter != "Tất cả" and emp.department != dept_filter:
                continue
            
            # Filter by role
            if role_filter != "Tất cả":
                if role_filter == "Bác sĩ" and emp.role != "doctor":
                    continue
                if role_filter == "Điều dưỡng" and emp.role != "nurse":
                    continue
            
            role_text = "Bác sĩ" if emp.role == "doctor" else "Điều dưỡng"
            self.employee_tree.insert("", "end", values=(
                emp.id,
                emp.name,
                role_text,
                emp.department,
                f"{emp.years_exp} năm",
                len(emp.days_off)
            ))
            count += 1
        
        self.emp_count_label.config(text=f"Hiển thị: {count}/{len(self.employees)} nhân viên")
    
    def on_employee_select(self, event):
        """Khi chọn nhân viên"""
        selection = self.employee_tree.selection()
        if not selection:
            return
        
        item = self.employee_tree.item(selection[0])
        emp_id = int(item['values'][0])
        
        # Find employee
        emp = next((e for e in self.employees if e.id == emp_id), None)
        if emp:
            self.selected_emp_id = emp_id
            role = "Bác sĩ" if emp.role == "doctor" else "Điều dưỡng"
            self.selected_emp_info.set(
                f"🔹 {emp.name} - {role} - {emp.department} - {emp.years_exp} năm kinh nghiệm"
            )
            
            # Load current days off
            self.selected_days_set = set(emp.days_off)
            self.draw_calendar()
            self.update_dayoff_summary()
    
    def draw_calendar(self):
        """Vẽ calendar"""
        from datetime import datetime, timedelta
        
        # Clear existing
        for widget in self.calendar_frame.winfo_children():
            widget.destroy()
        
        self.day_buttons.clear()
        
        # Update month label
        self.month_label_var.set(
            f"Tháng {self.current_date.month:02d}/{self.current_date.year}"
        )
        
        # Create calendar grid
        # Day headers
        days_header = ["T2", "T3", "T4", "T5", "T6", "T7", "CN"]
        for i, day_name in enumerate(days_header):
            label = tk.Label(self.calendar_frame, text=day_name,
                           font=('Arial', 10, 'bold'),
                           bg='#4472C4', fg='white',
                           width=8, height=1)
            label.grid(row=0, column=i, sticky="nsew", padx=1, pady=1)
        
        # Get first day and number of days in month
        first_day = self.current_date.replace(day=1)
        weekday = first_day.weekday()  # 0 = Monday
        
        # Calculate days in month
        if self.current_date.month == 12:
            next_month = first_day.replace(year=first_day.year + 1, month=1)
        else:
            next_month = first_day.replace(month=first_day.month + 1)
        
        days_in_month = (next_month - first_day).days
        
        # Draw days
        row = 1
        col = weekday
        
        for day in range(1, days_in_month + 1):
            day_index = day - 1  # 0-indexed
            is_selected = day_index in self.selected_days_set
            
            # Determine color
            if is_selected:
                bg_color = '#FF6B6B'  # Red for selected
                fg_color = 'white'
            else:
                bg_color = 'white'
                fg_color = 'black'
            
            btn = tk.Button(self.calendar_frame, text=str(day),
                          font=('Arial', 10),
                          width=8, height=3,
                          bg=bg_color, fg=fg_color,
                          relief='raised',
                          command=lambda d=day_index: self.toggle_day(d))
            btn.grid(row=row, column=col, sticky="nsew", padx=1, pady=1)
            
            self.day_buttons[day_index] = btn
            
            col += 1
            if col > 6:
                col = 0
                row += 1
        
        # Configure grid weights
        for i in range(7):
            self.calendar_frame.columnconfigure(i, weight=1)
        for i in range(row + 1):
            self.calendar_frame.rowconfigure(i, weight=1)
    
    def toggle_day(self, day_index):
        """Chọn/bỏ chọn ngày"""
        if not self.selected_emp_id:
            messagebox.showwarning("⚠️ Cảnh báo", 
                                  "Vui lòng chọn nhân viên trước!")
            return
        
        if day_index in self.selected_days_set:
            self.selected_days_set.remove(day_index)
            self.day_buttons[day_index].config(bg='white', fg='black')
        else:
            self.selected_days_set.add(day_index)
            self.day_buttons[day_index].config(bg='#FF6B6B', fg='white')
        
        self.update_dayoff_summary()
    
    def update_dayoff_summary(self):
        """Cập nhật thống kê ngày nghỉ"""
        count = len(self.selected_days_set)
        self.dayoff_summary.set(f"Đã chọn: {count} ngày")
    
    def change_month(self, delta):
        """Thay đổi tháng"""
        from datetime import datetime
        
        new_month = self.current_date.month + delta
        new_year = self.current_date.year
        
        if new_month > 12:
            new_month = 1
            new_year += 1
        elif new_month < 1:
            new_month = 12
            new_year -= 1
        
        self.current_date = self.current_date.replace(year=new_year, month=new_month, day=1)
        self.draw_calendar()
    
    def go_to_today(self):
        """Về tháng hiện tại"""
        from datetime import datetime
        self.current_date = datetime.now()
        self.draw_calendar()
    
    def save_dayoffs(self):
        """Lưu ngày nghỉ cho nhân viên"""
        if not self.selected_emp_id:
            messagebox.showwarning("⚠️ Cảnh báo", 
                                  "Vui lòng chọn nhân viên!")
            return
        
        emp = next((e for e in self.employees if e.id == self.selected_emp_id), None)
        if emp:
            emp.days_off = set(self.selected_days_set)
            
            # Update tree
            self.update_employee_list()
            self.filter_employees()
            
            messagebox.showinfo("✅ Thành công",
                              f"Đã lưu {len(self.selected_days_set)} ngày nghỉ cho:\n"
                              f"{emp.name}")
    
    def clear_all_dayoffs(self):
        """Xóa tất cả ngày nghỉ đã chọn"""
        if not self.selected_emp_id:
            messagebox.showwarning("⚠️ Cảnh báo",
                                  "Vui lòng chọn nhân viên!")
            return
        
        if messagebox.askyesno("⚠️ Xác nhận",
                              "Xóa tất cả ngày nghỉ đã chọn?"):
            self.selected_days_set.clear()
            self.draw_calendar()
            self.update_dayoff_summary()
    
    def show_dayoff_stats(self):
        """Hiển thị thống kê ngày nghỉ"""
        if not self.employees:
            messagebox.showwarning("⚠️ Cảnh báo",
                                  "Chưa có dữ liệu nhân viên!")
            return
        
        # Tính toán thống kê
        total_dayoffs = sum(len(e.days_off) for e in self.employees)
        avg_dayoffs = total_dayoffs / len(self.employees) if self.employees else 0
        
        # Nhân viên có nhiều ngày nghỉ nhất
        max_emp = max(self.employees, key=lambda e: len(e.days_off))
        min_emp = min(self.employees, key=lambda e: len(e.days_off))
        
        # Thống kê theo khoa
        dept_stats = {}
        for emp in self.employees:
            if emp.department not in dept_stats:
                dept_stats[emp.department] = []
            dept_stats[emp.department].append(len(emp.days_off))
        
        stats_text = f"""📊 THỐNG KÊ NGÀY NGHỈ

📈 Tổng quan:
   • Tổng nhân viên: {len(self.employees)}
   • Tổng ngày nghỉ: {total_dayoffs}
   • Trung bình: {avg_dayoffs:.1f} ngày/người

👤 Cực trị:
   • Nhiều nhất: {max_emp.name} - {len(max_emp.days_off)} ngày
   • Ít nhất: {min_emp.name} - {len(min_emp.days_off)} ngày

🏥 Theo khoa:"""
        
        for dept, dayoffs in dept_stats.items():
            avg = sum(dayoffs) / len(dayoffs)
            stats_text += f"\n   • {dept}: TB {avg:.1f} ngày"
        
        messagebox.showinfo("📊 Thống kê ngày nghỉ", stats_text)
    
    def setup_tab3_run(self):
        """Tab 3: Chạy và theo dõi"""
        # Main container
        main_frame = ttk.Frame(self.tab3)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # ===== CONTROL PANEL =====
        control_frame = ttk.LabelFrame(main_frame, text="🎮 Điều khiển", padding="10")
        control_frame.pack(fill="x", pady=(0, 10))
        
        # Buttons row 1
        btn_row1 = ttk.Frame(control_frame)
        btn_row1.pack(fill="x", pady=5)
        
        self.run_button = ttk.Button(btn_row1, text="▶️ Bắt đầu chạy",
                                     command=self.start_ga,
                                     style='Action.TButton',
                                     width=20)
        self.run_button.pack(side="left", padx=5)
        
        self.stop_button = ttk.Button(btn_row1, text="⏸️ Dừng lại",
                                      command=self.stop_ga,
                                      width=20,
                                      state="disabled")
        self.stop_button.pack(side="left", padx=5)
        
        self.clear_button = ttk.Button(btn_row1, text="🗑️ Xóa console",
                                       command=self.clear_console,
                                       width=15)
        self.clear_button.pack(side="left", padx=5)
        
        # Progress bar and status
        status_frame = ttk.Frame(control_frame)
        status_frame.pack(fill="x", pady=5)
        
        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(status_frame, 
                                           variable=self.progress_var,
                                           maximum=100,
                                           mode='determinate')
        self.progress_bar.pack(fill="x", pady=5)
        
        self.status_label = ttk.Label(status_frame, 
                                      text="Trạng thái: Chưa chạy",
                                      font=('Arial', 10, 'bold'),
                                      foreground='gray')
        self.status_label.pack(fill="x")
        
        # Statistics frame
        stats_frame = ttk.Frame(control_frame)
        stats_frame.pack(fill="x", pady=5)
        
        self.gen_label = ttk.Label(stats_frame, text="Thế hệ: 0/0", font=('Arial', 9))
        self.gen_label.pack(side="left", padx=10)
        
        self.fitness_label = ttk.Label(stats_frame, text="Fitness tốt nhất: -", font=('Arial', 9))
        self.fitness_label.pack(side="left", padx=10)
        
        self.time_label = ttk.Label(stats_frame, text="Thời gian: 0s", font=('Arial', 9))
        self.time_label.pack(side="left", padx=10)
        
        # ===== MAIN CONTENT: Console + Chart =====
        content_paned = ttk.PanedWindow(main_frame, orient=tk.VERTICAL)
        content_paned.pack(fill="both", expand=True)
        
        # Console panel
        console_frame = ttk.LabelFrame(content_paned, text="📟 Console Output")
        content_paned.add(console_frame, weight=1)
        
        # Console text with scrollbar
        console_container = ttk.Frame(console_frame)
        console_container.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.console_text = scrolledtext.ScrolledText(console_container,
                                                      wrap=tk.WORD,
                                                      height=15,
                                                      font=('Consolas', 9),
                                                      bg='#1E1E1E',
                                                      fg='#D4D4D4')
        self.console_text.pack(fill="both", expand=True)
        
        # Configure tags for colored output
        self.console_text.tag_config('info', foreground='#4EC9B0')
        self.console_text.tag_config('success', foreground='#4CAF50')
        self.console_text.tag_config('warning', foreground='#FFA500')
        self.console_text.tag_config('error', foreground='#F44336')
        self.console_text.tag_config('header', foreground='#FFD700', font=('Consolas', 9, 'bold'))
        
        # Chart panel
        chart_frame = ttk.LabelFrame(content_paned, text="📊 Biểu đồ hội tụ")
        content_paned.add(chart_frame, weight=1)
        
        # Create matplotlib figure
        self.fig = Figure(figsize=(8, 4), dpi=100, facecolor='white')
        self.ax = self.fig.add_subplot(111)
        self.ax.set_xlabel("Thế hệ (Generation)", fontsize=10)
        self.ax.set_ylabel("Fitness (Penalty)", fontsize=10)
        self.ax.set_title("Quá trình hội tụ của thuật toán GA", fontsize=12, fontweight='bold')
        self.ax.grid(True, alpha=0.3)
        
        # Canvas for matplotlib
        self.canvas = FigureCanvasTkAgg(self.fig, chart_frame)
        self.canvas.get_tk_widget().pack(fill="both", expand=True, padx=5, pady=5)
        
        # Initial message
        self.log_console("🎯 Hệ thống sẵn sàng. Nhấn 'Bắt đầu chạy' để khởi động thuật toán GA.\n", 'info')
        self.log_console("⚠️ Lưu ý: Đảm bảo đã tạo dữ liệu mẫu ở tab Cấu hình trước khi chạy!\n\n", 'warning')
    
    def log_console(self, text, tag='info'):
        """Ghi log vào console với màu sắc"""
        self.output_queue.put(('console', text, tag))
    
    def clear_console(self):
        """Xóa nội dung console"""
        self.console_text.delete(1.0, tk.END)
        self.log_console("Console đã được xóa.\n\n", 'info')
    
    def start_ga(self):
        """Bắt đầu chạy GA"""
        # Validate data
        if not self.employees:
            messagebox.showerror("❌ Lỗi",
                               "Chưa có dữ liệu nhân viên!\n\n"
                               "Vui lòng tạo dữ liệu mẫu ở tab Cấu hình trước.")
            return
        
        if self.is_running:
            messagebox.showwarning("⚠️ Cảnh báo",
                                  "Thuật toán đang chạy!")
            return
        
        # Confirm
        if not messagebox.askyesno("🚀 Xác nhận",
                                   f"Bắt đầu chạy thuật toán GA?\n\n"
                                   f"Cấu hình:\n"
                                   f"  • Số thế hệ: {self.config['GENERATIONS']}\n"
                                   f"  • Kích thước quần thể: {self.config['POPULATION_SIZE']}\n"
                                   f"  • Số nhân viên: {len(self.employees)}\n"
                                   f"  • Số ngày: {self.config['NUM_DAYS']}\n\n"
                                   f"Quá trình này có thể mất vài phút..."):
            return
        
        # Reset
        self.is_running = True
        self.history = []
        self.best_schedule = None
        
        # Update UI
        self.run_button.config(state="disabled")
        self.stop_button.config(state="normal")
        self.status_label.config(text="Trạng thái: Đang chạy...", foreground='blue')
        self.progress_var.set(0)
        
        # Clear chart
        self.ax.clear()
        self.ax.set_xlabel("Thế hệ (Generation)", fontsize=10)
        self.ax.set_ylabel("Fitness (Penalty)", fontsize=10)
        self.ax.set_title("Quá trình hội tụ của thuật toán GA", fontsize=12, fontweight='bold')
        self.ax.grid(True, alpha=0.3)
        self.canvas.draw()
        
        # Start thread
        thread = threading.Thread(target=self.run_ga_algorithm, daemon=True)
        thread.start()
    
    def stop_ga(self):
        """Dừng GA"""
        if messagebox.askyesno("⚠️ Xác nhận",
                              "Dừng thuật toán?\n\n"
                              "Tiến trình hiện tại sẽ bị hủy."):
            self.is_running = False
            self.run_button.config(state="normal")
            self.stop_button.config(state="disabled")
            self.status_label.config(text="Trạng thái: Đã dừng", foreground='orange')
            self.log_console("\n⏸️ Người dùng đã dừng thuật toán.\n\n", 'warning')
    
    def run_ga_algorithm(self):
        """Chạy thuật toán GA trong thread riêng"""
        import time
        start_time = time.time()
        
        try:
            self.log_console("=" * 80 + "\n", 'header')
            self.log_console("🚀 BẮT ĐẦU CHẠY THUẬT TOÁN DI TRUYỀN (GA)\n", 'header')
            self.log_console("=" * 80 + "\n\n", 'header')
            
            self.log_console(f"📋 Thông tin cấu hình:\n", 'info')
            self.log_console(f"   • Số thế hệ: {self.config['GENERATIONS']}\n", 'info')
            self.log_console(f"   • Kích thước quần thể: {self.config['POPULATION_SIZE']}\n", 'info')
            self.log_console(f"   • Số nhân viên: {len(self.employees)}\n", 'info')
            self.log_console(f"   • Số ngày lập lịch: {self.config['NUM_DAYS']}\n\n", 'info')
            
            # Tạo quần thể ban đầu
            self.log_console("🧬 Đang tạo quần thể ban đầu...\n", 'info')
            population = []
            for i in range(self.config['POPULATION_SIZE']):
                if not self.is_running:
                    return
                ind = ga_module.create_individual(self.employees, self.dept_to_rooms,
                                                  self.shifts, self.days)
                population.append(ind)
                if (i + 1) % 20 == 0:
                    self.log_console(f"   Đã tạo {i + 1}/{self.config['POPULATION_SIZE']} cá thể\n", 'info')
            
            self.log_console("✅ Hoàn thành tạo quần thể!\n\n", 'success')
            
            best_fit = float("inf")
            stagnation = 0
            
            self.log_console("🔄 Bắt đầu tiến hóa...\n\n", 'info')
            
            # Main GA loop
            for gen in range(self.config['GENERATIONS']):
                if not self.is_running:
                    self.log_console("\n⏸️ Thuật toán đã bị dừng.\n", 'warning')
                    return
                
                # Evaluate fitness
                scored = []
                for ind in population:
                    fit = ga_module.fitness(ind, self.employees, self.dept_to_rooms,
                                           self.shifts, self.days)
                    scored.append((fit, ind))
                scored.sort(key=lambda x: x[0])
                
                # Get best
                best = scored[0][1]
                fit = scored[0][0]
                
                self.history.append(fit)
                
                # Log progress
                if gen % 10 == 0 or gen == self.config['GENERATIONS'] - 1:
                    elapsed = time.time() - start_time
                    self.log_console(
                        f"Gen {gen + 1:3d}/{self.config['GENERATIONS']} | "
                        f"Fitness = {fit:,.0f} | "
                        f"Time: {elapsed:.1f}s\n",
                        'info'
                    )
                
                # Update UI
                progress = ((gen + 1) / self.config['GENERATIONS']) * 100
                self.output_queue.put(('progress', progress, gen + 1, fit, elapsed))
                
                # Update chart every 5 generations
                if gen % 5 == 0 or gen == self.config['GENERATIONS'] - 1:
                    self.output_queue.put(('chart', None))
                
                # Check improvement
                if fit < best_fit:
                    best_fit = fit
                    stagnation = 0
                    self.best_schedule = copy.deepcopy(best)
                else:
                    stagnation += 1
                
                # Hill climbing if stagnated
                if stagnation >= self.config['STAGNATION_LIMIT']:
                    self.log_console(f"   🔧 Hill Climbing triggered at Gen {gen + 1}\n", 'warning')
                    best = ga_module.hill_climb(best, self.employees, self.dept_to_rooms,
                                                self.shifts, self.days,
                                                self.config['HILL_CLIMB_STEPS'])
                    stagnation = 0
                
                # Create new population
                new_pop = [copy.deepcopy(scored[i][1]) for i in range(self.config['ELITE_SIZE'])]
                
                while len(new_pop) < self.config['POPULATION_SIZE']:
                    p1 = ga_module.tournament_selection(scored)
                    p2 = ga_module.tournament_selection(scored)
                    child = ga_module.crossover_uniform(p1, p2, self.employees, self.dept_to_rooms)
                    child = ga_module.mutate_scramble(child, self.employees, self.dept_to_rooms,
                                                      self.shifts, self.days,
                                                      self.config['MUTATION_RATE'])
                    child = ga_module.mutate_balance_hours(child, self.employees, self.dept_to_rooms,
                                                          self.shifts, self.days, 0.3)
                    new_pop.append(child)
                
                population = new_pop
            
            # Kết thúc
            if self.is_running:
                elapsed = time.time() - start_time
                
                self.log_console("\n" + "=" * 80 + "\n", 'header')
                self.log_console("✅ HOÀN THÀNH THUẬT TOÁN!\n", 'success')
                self.log_console("=" * 80 + "\n\n", 'header')
                
                self.log_console(f"📊 Kết quả:\n", 'success')
                self.log_console(f"   • Fitness tốt nhất: {best_fit:,.0f}\n", 'success')
                self.log_console(f"   • Thời gian chạy: {elapsed:.1f}s ({elapsed/60:.1f} phút)\n", 'success')
                self.log_console(f"   • Số thế hệ: {self.config['GENERATIONS']}\n\n", 'success')
                
                # Convert schedule to dashboard format
                self.best_schedule = self.convert_schedule_format(self.best_schedule)
                
                # Kiểm tra ràng buộc
                self.log_console("🔍 Đang kiểm tra ràng buộc...\n", 'info')
                hard_violations, soft_violations, soft_metrics, soft_stats = \
                    ga_module.check_constraints_detailed(best, self.employees,
                                                        self.dept_to_rooms, self.shifts, self.days)
                
                total_hard = sum(len(v) for v in hard_violations.values())
                total_soft = sum(len(v) for v in soft_violations.values())
                
                self.log_console(f"\n📈 Thống kê vi phạm:\n", 'info')
                if total_hard == 0:
                    self.log_console(f"   ✅ Vi phạm ràng buộc cứng: 0 (HOÀN HẢO!)\n", 'success')
                else:
                    self.log_console(f"   ⚠️ Vi phạm ràng buộc cứng: {total_hard}\n", 'warning')
                
                self.log_console(f"   📊 Vi phạm ràng buộc mềm: {total_soft}\n", 'info')
                self.log_console(f"   ⏰ Giờ làm trung bình: {soft_metrics['avg_hours']:.1f}h\n", 'info')
                self.log_console(f"   📅 Số ca trực trung bình: {soft_metrics['avg_shifts']:.1f} ca\n\n", 'info')
                
                self.log_console("🎉 Bạn có thể xem kết quả chi tiết ở tab Dashboard!\n", 'success')
                
                self.output_queue.put(('complete', elapsed, best_fit))
        
        except Exception as e:
            self.log_console(f"\n❌ LỖI: {str(e)}\n", 'error')
            self.log_console(f"Chi tiết: {type(e).__name__}\n\n", 'error')
            self.output_queue.put(('error', str(e)))
        
        finally:
            self.is_running = False
            self.run_button.config(state="normal")
            self.stop_button.config(state="disabled")
    
    def check_queue(self):
        """Kiểm tra queue để cập nhật UI từ thread"""
        try:
            while True:
                msg = self.output_queue.get_nowait()
                msg_type = msg[0]
                
                if msg_type == 'console':
                    _, text, tag = msg
                    self.console_text.insert(tk.END, text, tag)
                    self.console_text.see(tk.END)
                
                elif msg_type == 'progress':
                    _, progress, gen, fit, elapsed = msg
                    self.progress_var.set(progress)
                    self.gen_label.config(text=f"Thế hệ: {gen}/{self.config['GENERATIONS']}")
                    self.fitness_label.config(text=f"Fitness tốt nhất: {fit:,.0f}")
                    self.time_label.config(text=f"Thời gian: {elapsed:.1f}s")
                
                elif msg_type == 'chart':
                    self.update_chart()
                
                elif msg_type == 'complete':
                    _, elapsed, best_fit = msg
                    self.status_label.config(text=f"Trạng thái: Hoàn thành ({elapsed:.1f}s)", 
                                           foreground='green')
                    self.progress_var.set(100)
                
                elif msg_type == 'error':
                    _, error_msg = msg
                    self.status_label.config(text=f"Trạng thái: Lỗi", foreground='red')
                    messagebox.showerror("❌ Lỗi", f"Có lỗi xảy ra:\n\n{error_msg}")
        
        except queue.Empty:
            pass
        
        finally:
            self.root.after(100, self.check_queue)
    
    def update_chart(self):
        """Cập nhật biểu đồ hội tụ"""
        if not self.history:
            return
        
        self.ax.clear()
        self.ax.plot(range(1, len(self.history) + 1), self.history,
                    'b-', linewidth=2, label='Fitness')
        
        # Thêm đường trung bình
        if len(self.history) > 10:
            import numpy as np
            window = 10
            moving_avg = np.convolve(self.history, np.ones(window)/window, mode='valid')
            self.ax.plot(range(window, len(self.history) + 1), moving_avg,
                        'r--', linewidth=1.5, alpha=0.7, label='Trung bình động')
        
        self.ax.set_xlabel("Thế hệ (Generation)", fontsize=10)
        self.ax.set_ylabel("Fitness (Penalty)", fontsize=10)
        self.ax.set_title("Quá trình hội tụ của thuật toán GA", fontsize=12, fontweight='bold')
        self.ax.grid(True, alpha=0.3)
        self.ax.legend(loc='upper right')
        
        # Format y-axis với dấu phẩy
        self.ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: f'{int(x):,}'))
        
        self.fig.tight_layout()
        self.canvas.draw()
    
    def convert_schedule_format(self, ga_schedule):
        """Chuyển đổi schedule từ GA format sang dashboard format"""
        # GA format: schedule[day][shift_name][room] = [emp_ids]
        # Dashboard format: schedule[day] = [{'employee': emp, 'shift': shift, 'room': room}]
        
        emp_dict = {e.id: e for e in self.employees}
        shift_dict = {s.name: s for s in self.shifts}
        
        # Create a simple Room class for storing room info
        class SimpleRoom:
            def __init__(self, name):
                self.name = name
        
        converted = {}
        for day in ga_schedule:
            converted[day] = []
            for shift_name in ga_schedule[day]:
                for room_name in ga_schedule[day][shift_name]:
                    emp_ids = ga_schedule[day][shift_name][room_name]
                    for emp_id in emp_ids:
                        if emp_id in emp_dict:
                            converted[day].append({
                                'employee': emp_dict[emp_id],
                                'shift': shift_dict[shift_name],
                                'room': SimpleRoom(room_name)
                            })
        
        return converted
        
        return converted
    
    def setup_tab4_dashboard(self):
        """Tab 4: Dashboard và xuất Excel"""
        # Main container
        main_frame = ttk.Frame(self.tab4)
        main_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # ===== CONTROL PANEL =====
        control_frame = ttk.LabelFrame(main_frame, text="🎛️ Điều khiển", padding="10")
        control_frame.pack(fill="x", pady=(0, 10))
        
        # Row 1: Filters
        filter_row = ttk.Frame(control_frame)
        filter_row.pack(fill="x", pady=5)
        
        ttk.Label(filter_row, text="Khoa:").pack(side="left", padx=(0, 5))
        self.dashboard_dept_var = tk.StringVar(value="Tất cả")
        self.dashboard_dept_combo = ttk.Combobox(filter_row, 
                                           textvariable=self.dashboard_dept_var,
                                           state="readonly",
                                           width=20)
        self.dashboard_dept_combo['values'] = ["Tất cả"] + self.config['DEPARTMENTS']
        self.dashboard_dept_combo.pack(side="left", padx=5)
        self.dashboard_dept_combo.bind('<<ComboboxSelected>>', lambda e: self.refresh_dashboard())
        
        ttk.Label(filter_row, text="Nhân viên:").pack(side="left", padx=(20, 5))
        self.dashboard_emp_var = tk.StringVar(value="Tất cả")
        self.dashboard_emp_combo = ttk.Combobox(filter_row,
                                               textvariable=self.dashboard_emp_var,
                                               state="readonly",
                                               width=25)
        self.dashboard_emp_combo.pack(side="left", padx=5)
        self.dashboard_emp_combo.bind('<<ComboboxSelected>>', lambda e: self.refresh_dashboard())
        
        ttk.Button(filter_row, text="🔄 Làm mới",
                  command=self.refresh_dashboard,
                  width=12).pack(side="left", padx=10)
        
        # Row 2: Export buttons
        export_row = ttk.Frame(control_frame)
        export_row.pack(fill="x", pady=5)
        
        ttk.Button(export_row, text="📊 Xuất Excel (Toàn bộ)",
                  command=lambda: self.export_to_excel('all'),
                  style='Action.TButton',
                  width=25).pack(side="left", padx=5)
        
        ttk.Button(export_row, text="📋 Xuất Excel (Theo khoa)",
                  command=lambda: self.export_to_excel('department'),
                  width=25).pack(side="left", padx=5)
        
        ttk.Button(export_row, text="👤 Xuất Excel (Theo NV)",
                  command=lambda: self.export_to_excel('employee'),
                  width=25).pack(side="left", padx=5)
        
        # ===== STATISTICS PANEL =====
        stats_frame = ttk.LabelFrame(main_frame, text="📈 Thống kê tổng quan", padding="10")
        stats_frame.pack(fill="x", pady=(0, 10))
        
        # Create stats labels
        stats_container = ttk.Frame(stats_frame)
        stats_container.pack(fill="x")
        
        self.total_shifts_label = ttk.Label(stats_container, 
                                           text="Tổng ca trực: -",
                                           font=('Arial', 10, 'bold'))
        self.total_shifts_label.pack(side="left", padx=20)
        
        self.avg_hours_label = ttk.Label(stats_container,
                                         text="Giờ làm TB: -",
                                         font=('Arial', 10, 'bold'))
        self.avg_hours_label.pack(side="left", padx=20)
        
        self.violations_label = ttk.Label(stats_container,
                                         text="Vi phạm: -",
                                         font=('Arial', 10, 'bold'))
        self.violations_label.pack(side="left", padx=20)
        
        self.fitness_dashboard_label = ttk.Label(stats_container,
                                                text="Fitness: -",
                                                font=('Arial', 10, 'bold'))
        self.fitness_dashboard_label.pack(side="left", padx=20)
        
        # ===== CALENDAR VIEW =====
        calendar_frame = ttk.LabelFrame(main_frame, text="📅 Lịch trực", padding="5")
        calendar_frame.pack(fill="both", expand=True)
        
        # Create canvas with scrollbar
        canvas_container = ttk.Frame(calendar_frame)
        canvas_container.pack(fill="both", expand=True)
        
        self.dashboard_canvas = tk.Canvas(canvas_container, bg='white')
        v_scrollbar = ttk.Scrollbar(canvas_container, orient="vertical",
                                   command=self.dashboard_canvas.yview)
        h_scrollbar = ttk.Scrollbar(canvas_container, orient="horizontal",
                                   command=self.dashboard_canvas.xview)
        
        self.dashboard_canvas.configure(yscrollcommand=v_scrollbar.set,
                                       xscrollcommand=h_scrollbar.set)
        
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        self.dashboard_canvas.pack(side="left", fill="both", expand=True)
        
        # Frame inside canvas
        self.dashboard_inner_frame = ttk.Frame(self.dashboard_canvas)
        self.canvas_window = self.dashboard_canvas.create_window((0, 0),
                                                                 window=self.dashboard_inner_frame,
                                                                 anchor="nw")
        
        # Configure scroll region
        self.dashboard_inner_frame.bind('<Configure>',
                                       lambda e: self.dashboard_canvas.configure(
                                           scrollregion=self.dashboard_canvas.bbox("all")))
        
        # Mouse wheel scroll for dashboard
        def _on_dashboard_mousewheel(event):
            self.dashboard_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        def _on_dashboard_h_mousewheel(event):
            self.dashboard_canvas.xview_scroll(int(-1*(event.delta/120)), "units")
        
        self.dashboard_canvas.bind("<MouseWheel>", _on_dashboard_mousewheel)
        self.dashboard_canvas.bind("<Shift-MouseWheel>", _on_dashboard_h_mousewheel)
        
        # Initial message
        ttk.Label(self.dashboard_inner_frame,
                 text="Chưa có dữ liệu lịch trực.\n\n"
                      "Vui lòng chạy thuật toán GA ở tab 'Chạy và theo dõi' trước.",
                 font=('Arial', 12),
                 foreground='gray').pack(pady=50)
    
    def refresh_dashboard(self):
        """Làm mới dashboard với dữ liệu mới"""
        if not self.best_schedule:
            messagebox.showinfo("ℹ️ Thông báo",
                               "Chưa có dữ liệu lịch trực!\n\n"
                               "Vui lòng chạy thuật toán GA ở tab 'Chạy và theo dõi' trước.")
            return
        
        # Update employee combo based on department
        dept = self.dashboard_dept_var.get()
        if dept == "Tất cả":
            emp_list = ["Tất cả"] + [e.name for e in self.employees]
        else:
            emp_list = ["Tất cả"] + [e.name for e in self.employees if e.department == dept]
        
        self.dashboard_emp_combo['values'] = emp_list
        if self.dashboard_emp_var.get() not in emp_list:
            self.dashboard_emp_var.set("Tất cả")
        
        # Clear existing widgets
        for widget in self.dashboard_inner_frame.winfo_children():
            widget.destroy()
        
        # Filter schedule
        filtered_schedule = self.filter_schedule()
        
        # Update statistics
        self.update_dashboard_stats(filtered_schedule)
        
        # Draw calendar
        self.draw_dashboard_calendar(filtered_schedule)
    
    def filter_schedule(self):
        """Lọc lịch trực theo bộ lọc"""
        dept = self.dashboard_dept_var.get()
        emp_name = self.dashboard_emp_var.get()
        
        filtered = {}
        for day, day_schedule in self.best_schedule.items():
            filtered[day] = []
            for shift_data in day_schedule:
                emp = shift_data['employee']
                
                # Filter by department
                if dept != "Tất cả" and emp.department != dept:
                    continue
                
                # Filter by employee
                if emp_name != "Tất cả" and emp.name != emp_name:
                    continue
                
                filtered[day].append(shift_data)
        
        return filtered
    
    def update_dashboard_stats(self, schedule):
        """Cập nhật thống kê dashboard"""
        if not schedule:
            return
        
        # Count total shifts
        total_shifts = sum(len(shifts) for shifts in schedule.values())
        self.total_shifts_label.config(text=f"Tổng ca trực: {total_shifts}")
        
        # Calculate average hours
        employee_hours = {}
        for day_schedule in schedule.values():
            for shift_data in day_schedule:
                emp = shift_data['employee']
                shift = shift_data['shift']
                if emp.name not in employee_hours:
                    employee_hours[emp.name] = 0
                employee_hours[emp.name] += shift.hours
        
        if employee_hours:
            avg_hours = sum(employee_hours.values()) / len(employee_hours)
            self.avg_hours_label.config(text=f"Giờ làm TB: {avg_hours:.1f}h")
        else:
            self.avg_hours_label.config(text="Giờ làm TB: -")
        
        # Check violations (only for full schedule)
        if self.dashboard_dept_var.get() == "Tất cả" and self.dashboard_emp_var.get() == "Tất cả":
            # Need to convert back to GA format for validation
            ga_format_schedule = self.convert_to_ga_format(self.best_schedule)
            
            hard_violations, soft_violations, _, _ = \
                ga_module.check_constraints_detailed(ga_format_schedule, self.employees,
                                                    self.dept_to_rooms, self.shifts, self.days)
            
            total_hard = sum(len(v) for v in hard_violations.values())
            total_soft = sum(len(v) for v in soft_violations.values())
            
            if total_hard == 0:
                self.violations_label.config(text=f"✅ Vi phạm: {total_soft} (mềm)",
                                           foreground='green')
            else:
                self.violations_label.config(text=f"⚠️ Vi phạm: {total_hard} cứng, {total_soft} mềm",
                                           foreground='red')
            
            # Calculate fitness
            fitness = ga_module.fitness(ga_format_schedule, self.employees,
                                       self.dept_to_rooms, self.shifts, self.days)
            self.fitness_dashboard_label.config(text=f"Fitness: {fitness:,.0f}")
        else:
            self.violations_label.config(text="Vi phạm: -", foreground='black')
            self.fitness_dashboard_label.config(text="Fitness: -")
    
    def convert_to_ga_format(self, dashboard_schedule):
        """Chuyển đổi từ dashboard format về GA format"""
        # Dashboard format: schedule[day] = [{'employee': emp, 'shift': shift, 'room': room}]
        # GA format: schedule[day][shift_name][room] = [emp_ids]
        
        ga_schedule = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))
        
        for day, shift_list in dashboard_schedule.items():
            for item in shift_list:
                emp = item['employee']
                shift = item['shift']
                room = item['room']
                
                if emp.id not in ga_schedule[day][shift.name][room.name]:
                    ga_schedule[day][shift.name][room.name].append(emp.id)
        
        return dict(ga_schedule)
    
    def draw_dashboard_calendar(self, schedule):
        """Vẽ lịch trực dạng bảng"""
        if not schedule:
            ttk.Label(self.dashboard_inner_frame,
                     text="Không có dữ liệu phù hợp với bộ lọc.",
                     font=('Arial', 11),
                     foreground='gray').pack(pady=30)
            return
        
        # Colors for shifts
        shift_colors = {
            'Sáng': '#E3F2FD',    # Light blue
            'Chiều': '#FFF3E0',   # Light orange
            'Tối': '#F3E5F5',     # Light purple
            'Đêm': '#E8F5E9'      # Light green
        }
        
        # Create table frame
        table_frame = ttk.Frame(self.dashboard_inner_frame)
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Header row
        header_frame = ttk.Frame(table_frame)
        header_frame.grid(row=0, column=0, columnspan=len(self.days) + 1, sticky="ew")
        
        # Empty cell for row header
        header_label = tk.Label(header_frame, text="Ca trực",
                               font=('Arial', 10, 'bold'),
                               bg='#1976D2', fg='white',
                               width=15, height=2,
                               relief='solid', borderwidth=1)
        header_label.grid(row=0, column=0, sticky="nsew")
        
        # Day headers
        for col, day in enumerate(self.days, 1):
            day_label = tk.Label(header_frame, text=day,
                                font=('Arial', 10, 'bold'),
                                bg='#2196F3', fg='white',
                                width=20, height=2,
                                relief='solid', borderwidth=1)
            day_label.grid(row=0, column=col, sticky="nsew")
        
        # Data rows - grouped by shift
        row = 1
        for shift in self.shifts:
            # Shift name label
            shift_label = tk.Label(table_frame,
                                  text=f"{shift.name}\n({shift.start}-{shift.end})",
                                  font=('Arial', 9, 'bold'),
                                  bg='#90CAF9', fg='black',
                                  width=15,
                                  relief='solid', borderwidth=1,
                                  justify='center')
            shift_label.grid(row=row, column=0, sticky="nsew", pady=1)
            
            # Cells for each day
            for col, day in enumerate(self.days, 1):
                # Get employees for this shift and day
                employees_in_shift = [
                    sd for sd in schedule.get(day, [])
                    if sd['shift'].name == shift.name
                ]
                
                # Create cell content
                if employees_in_shift:
                    content = "\n".join([
                        f"• {sd['employee'].name} ({sd['room'].name})"
                        for sd in employees_in_shift
                    ])
                    bg_color = shift_colors.get(shift.name, 'white')
                else:
                    content = "-"
                    bg_color = 'white'
                
                cell = tk.Label(table_frame, text=content,
                               font=('Arial', 8),
                               bg=bg_color,
                               width=20,
                               relief='solid', borderwidth=1,
                               justify='left',
                               anchor='nw',
                               padx=5, pady=5)
                cell.grid(row=row, column=col, sticky="nsew", pady=1, padx=1)
            
            row += 1
    
    def export_to_excel(self, export_type):
        """Xuất lịch trực ra Excel"""
        if not self.best_schedule:
            messagebox.showwarning("⚠️ Cảnh báo",
                                  "Chưa có dữ liệu lịch trực!\n\n"
                                  "Vui lòng chạy thuật toán GA trước.")
            return
        
        # Choose filename
        default_name = f"Lich_truc_{export_type}"
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile=default_name
        )
        
        if not filename:
            return
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            if export_type == 'all':
                ws = wb.active
                ws.title = "Toàn bộ"
                self._write_schedule_to_sheet(ws, self.best_schedule, "Toàn bộ")
            
            elif export_type == 'department':
                wb.remove(wb.active)  # Remove default sheet
                for dept in self.config['DEPARTMENTS']:
                    ws = wb.create_sheet(title=dept[:31])  # Excel limit 31 chars
                    dept_schedule = self._filter_by_department(dept)
                    self._write_schedule_to_sheet(ws, dept_schedule, dept)
            
            elif export_type == 'employee':
                wb.remove(wb.active)
                for emp in self.employees:
                    ws = wb.create_sheet(title=emp.name[:31])
                    emp_schedule = self._filter_by_employee(emp.name)
                    self._write_schedule_to_sheet(ws, emp_schedule, emp.name)
            
            # Save workbook
            wb.save(filename)
            
            if messagebox.askyesno("✅ Thành công",
                                  f"Đã xuất file Excel thành công!\n\n"
                                  f"File: {filename}\n\n"
                                  f"Bạn có muốn mở file?"):
                import os
                os.startfile(filename)
        
        except Exception as e:
            messagebox.showerror("❌ Lỗi", f"Không thể xuất file Excel:\n\n{str(e)}")
    
    def _write_schedule_to_sheet(self, ws, schedule, title):
        """Ghi lịch trực vào sheet Excel"""
        # Convert to GA format first
        ga_schedule = self.convert_to_ga_format(schedule)
        emp_dict = {e.id: e for e in self.employees}
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
        shift_fill = PatternFill(start_color="90CAF9", end_color="90CAF9", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"LỊCH TRỰC - {title.upper()}"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = center_align
        ws.row_dimensions[1].height = 30
        
        # Headers
        ws['A3'] = "Ca trực"
        ws['A3'].font = header_font
        ws['A3'].fill = header_fill
        ws['A3'].alignment = center_align
        ws['A3'].border = border
        
        for col, day in enumerate(self.days, 2):
            cell = ws.cell(row=3, column=col)
            cell.value = day
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = border
        
        # Data
        row = 4
        for shift in self.shifts:
            # Shift name
            shift_cell = ws.cell(row=row, column=1)
            shift_cell.value = f"{shift.name}\n({shift.start_hour}-{shift.end_hour})"
            shift_cell.font = Font(bold=True)
            shift_cell.fill = shift_fill
            shift_cell.alignment = center_align
            shift_cell.border = border
            
            # Data for each day
            for col, day in enumerate(self.days, 2):
                employees_in_shift = [
                    sd for sd in schedule.get(day, [])
                    if sd['shift'].name == shift.name
                ]
                
                content = "\n".join([
                    f"{sd['employee'].name} ({sd['room'].name})"
                    for sd in employees_in_shift
                ]) if employees_in_shift else "-"
                
                cell = ws.cell(row=row, column=col)
                cell.value = content
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = border
            
            ws.row_dimensions[row].height = 60
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        for col in range(2, len(self.days) + 2):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 30
        
        # Add statistics at bottom
        row += 2
        stats_cell = ws.cell(row=row, column=1)
        stats_cell.value = "THỐNG KÊ:"
        stats_cell.font = Font(bold=True)
        
        row += 1
        total_shifts = sum(len(shifts) for shifts in schedule.values())
        ws.cell(row=row, column=1).value = f"Tổng số ca trực: {total_shifts}"
        
        row += 1
        employee_set = set()
        for day_schedule in schedule.values():
            for sd in day_schedule:
                employee_set.add(sd['employee'].name)
        ws.cell(row=row, column=1).value = f"Số nhân viên: {len(employee_set)}"
    
    def _filter_by_department(self, dept):
        """Lọc lịch theo khoa"""
        filtered = {}
        for day, day_schedule in self.best_schedule.items():
            filtered[day] = [
                sd for sd in day_schedule
                if sd['employee'].department == dept
            ]
        return filtered
    
    def _filter_by_employee(self, emp_name):
        """Lọc lịch theo nhân viên"""
        filtered = {}
        for day, day_schedule in self.best_schedule.items():
            filtered[day] = [
                sd for sd in day_schedule
                if sd['employee'].name == emp_name
            ]
        return filtered


def main():
    root = tk.Tk()
    app = HospitalScheduleApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
